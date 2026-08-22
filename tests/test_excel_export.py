import os
import tempfile
import unittest
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.excel_export import build_order_frames, export_excel


SAMPLE_BILLS = [
    {
        "trang_thai": "Chờ duyệt",
        "ma_phieu": "DHEK07423MW2022",
        "ma_kh": "KH001230",
        "ten_kh": "TP. Chị Xuân combo",
        "sdt": "01234567898",
        "ngay_dat": "2022-11-16T17:00:00.000Z",
        "tong_tien_hang": 18000,
        "phai_thanh_toan": "17000",
        "san_pham": [
            {
                "stt": 1,
                "ma_sp": "",
                "ten_sp": "Anginovag(Hôp 1lo 10ml)",
                "so_luong": "1",
                "don_gia": "0",
                "thanh_tien": 0,
                "is_km": False,
            },
            {
                "stt": 2,
                "ma_sp": "00008",
                "ten_sp": "Bio baby ăn ngon Vkgold (Hộp 20 gói)",
                "so_luong": "1",
                "is_km": True,
            },
        ],
    },
    {
        "trang_thai": "Chờ duyệt",
        "ma_phieu": "DHEK07422MW2022",
        "ma_kh": "KH001230",
        "ngay_dat": "2022-11-16T17:00:00.000Z",
        "phai_thanh_toan": "17000",
        "san_pham": [
            {"stt": 1, "ma_sp": "A001", "so_luong": "1", "is_km": False},
            {"stt": 2, "ma_sp": "00008", "so_luong": "1", "is_km": True},
        ],
    },
]


class BuildOrderFramesTests(unittest.TestCase):
    def test_sales_and_promotional_items_share_one_detail_table(self):
        header, detail = build_order_frames(SAMPLE_BILLS)

        self.assertEqual(len(header), 2)
        self.assertEqual(len(detail), 4)
        self.assertEqual(int(detail["is_km"].sum()), 2)
        self.assertEqual(
            set(detail.loc[detail["is_km"], "loai_hang"].dropna()),
            {"Khuyến mãi"},
        )

    def test_codes_numeric_values_and_timezone_are_normalized(self):
        header, detail = build_order_frames(SAMPLE_BILLS)

        promo = detail[(detail["ma_phieu"] == "DHEK07423MW2022") & (detail["stt"] == 2)].iloc[0]
        self.assertEqual(promo["ma_sp"], "00008")
        self.assertEqual(float(promo["so_luong"]), 1.0)
        self.assertEqual(float(header.iloc[0]["phai_thanh_toan"]), 17000.0)
        self.assertEqual(header.iloc[0]["ngay_dat"], pd.Timestamp("2022-11-17 00:00:00"))

    def test_duplicate_line_key_is_rejected(self):
        duplicate = [dict(SAMPLE_BILLS[0])]
        duplicate[0]["san_pham"] = [
            {"stt": 1, "ma_sp": "A", "is_km": False},
            {"stt": 1, "ma_sp": "B", "is_km": False},
        ]
        with self.assertRaisesRegex(ValueError, "duplicate key"):
            build_order_frames(duplicate)


class ExportExcelTests(unittest.TestCase):
    def test_order_workbook_has_two_analytics_friendly_sheets(self):
        previous_cwd = Path.cwd()
        with tempfile.TemporaryDirectory() as temp_dir:
            try:
                os.chdir(temp_dir)
                path = export_excel(SAMPLE_BILLS, "DonBanHang", date(2022, 11, 17), "order")
                workbook = load_workbook(path, data_only=True)
                self.assertEqual(workbook.sheetnames, ["DonHang", "ChiTietSP"])

                sheet = workbook["ChiTietSP"]
                headers = [cell.value for cell in sheet[1]]
                ma_sp_column = headers.index("ma_sp") + 1
                values = [
                    sheet.cell(row=row, column=ma_sp_column).value
                    for row in range(2, sheet.max_row + 1)
                ]
                self.assertIn("00008", values)
                self.assertTrue(
                    all(
                        sheet.cell(row=row, column=ma_sp_column).number_format == "@"
                        for row in range(2, sheet.max_row + 1)
                    )
                )
            finally:
                os.chdir(previous_cwd)


if __name__ == "__main__":
    unittest.main()
