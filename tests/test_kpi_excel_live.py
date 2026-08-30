from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

import openpyxl
import pandas as pd

from kpi.manual_labels import (
    DETAIL_SHEET,
    ManualLabelIndex,
    labels_from_sheet,
    load_manual_labels,
)
from kpi.workbook_formulas import (
    CUSTOMER_HEADERS,
    SUMMARY_HEADERS,
    replace_customer_rows,
    replace_summary_rows,
    update_customer_image_formulas,
    write_parameters,
)


class ManualLabelTests(unittest.TestCase):
    @staticmethod
    def _detail_sheet(label: str = "Bien_hieu", record_id: str = "rid-1"):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = DETAIL_SHEET
        sheet.cell(5, 2, "NV A")
        sheet.cell(5, 3, "2026-08-05")
        sheet.cell(5, 4, "KH001")
        sheet.cell(5, 6, "1")
        sheet.cell(5, 8, label)
        sheet.cell(5, 14, "Mở ảnh")
        sheet.cell(5, 14).hyperlink = "https://example.test/images/a.jpg"
        sheet.cell(5, 27, record_id)
        return workbook, sheet

    def test_labels_are_loaded_and_record_id_has_priority(self):
        workbook, sheet = self._detail_sheet()
        index = labels_from_sheet(sheet, Path("old.xlsx"))
        self.assertEqual(index.by_record_id["rid-1"], "Bien_hieu")
        self.assertEqual(
            index.lookup(
                {
                    "record_id": "rid-1",
                    "ten_nhan_vien": "other",
                    "ngay": "2026-08-30",
                    "ma_kh": "other",
                    "stt_hinh": 99,
                    "hinh_anh": "https://example.test/other.jpg",
                }
            ),
            "Bien_hieu",
        )
        workbook.close()

    def test_fallback_key_survives_url_query_wrapping(self):
        workbook, sheet = self._detail_sheet(record_id="")
        sheet.cell(5, 14).hyperlink = (
            "https://wrapper.test/open?url=https%3A%2F%2Fcdn.test%2Fimages%2Fa.jpg"
        )
        index = labels_from_sheet(sheet, Path("old.xlsx"))
        self.assertEqual(
            index.lookup(
                {
                    "ten_nhan_vien": " nv a ",
                    "ngay": "2026-08-06",
                    "ma_kh": "KH001",
                    "stt_hinh": "STT 1",
                    "hinh_anh": "https://cdn.test/images/a.jpg",
                }
            ),
            "Bien_hieu",
        )
        workbook.close()

    def test_overlay_prefers_newer_manual_label(self):
        old = ManualLabelIndex({"rid": "Bien_hieu"}, {}, {})
        new = ManualLabelIndex({"rid": "Trung_bay"}, {}, {})
        self.assertEqual(old.overlay(new).lookup({"record_id": "rid"}), "Trung_bay")

    def test_invalid_manual_label_fails_closed(self):
        workbook, sheet = self._detail_sheet(label="Tự do")
        with self.assertRaises(RuntimeError):
            labels_from_sheet(sheet, Path("bad.xlsx"))
        workbook.close()

    def test_load_manual_labels_round_trip_and_missing_file(self):
        self.assertEqual(load_manual_labels(Path("does-not-exist.xlsx")), ManualLabelIndex.empty())
        workbook, _ = self._detail_sheet()
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "prior.xlsx"
            workbook.save(path)
            loaded = load_manual_labels(path)
            self.assertEqual(loaded.lookup({"record_id": "rid-1"}), "Bien_hieu")
        workbook.close()


class LiveWorkbookFormulaTests(unittest.TestCase):
    @staticmethod
    def _make_sheet(workbook, title: str, column_count: int):
        sheet = workbook.create_sheet(title)
        for column in range(1, column_count + 1):
            sheet.cell(4, column, f"H{column}")
            sheet.cell(5, column, "")
        return sheet

    def test_parameters_customer_and_summary_are_live_formulas(self):
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        params = self._make_sheet(workbook, "Tham_so", 2)
        customer = self._make_sheet(workbook, "Chi_tiet_Khach_hang", max(CUSTOMER_HEADERS))
        summary = self._make_sheet(workbook, "Tong_hop_KPI_Nhan_vien", max(SUMMARY_HEADERS))

        write_parameters(params, pd.Timestamp("2026-08-01"), ("history warning",))
        self.assertTrue(str(params["B11"].value).startswith("=NETWORKDAYS.INTL"))
        self.assertEqual(params["B4"].value, 50)
        self.assertEqual(params["B7"].value, 30000)
        self.assertEqual(params["A15"].value, "history warning")

        facts = pd.DataFrame(
            [
                {
                    "ten_nhan_vien": "NV A",
                    "ma_kh": "KH001",
                    "ten_kh": "Cửa hàng A",
                    "visit_count_m": 2,
                    "first_activity_date": pd.Timestamp("2026-07-01"),
                    "max_order_2m_ktb": 3.2,
                    "total_order_2m_ktb": 4.0,
                    "order_count_2m": 2,
                    "ghi_ton_2m": True,
                    "valid_sign_note_2m": False,
                    "ghi_chu_2m": "",
                    "period_start": pd.Timestamp("2026-08-01"),
                }
            ]
        )
        customer_end = replace_customer_rows(customer, facts)
        self.assertEqual(customer_end, 5)
        self.assertIn("KHTC", customer["I5"].value)
        self.assertIn("Không Đạt", customer["P5"].value)
        self.assertEqual(customer["J5"].value, 1)

        update_customer_image_formulas(customer, detail_count=4)
        self.assertIn("Bien_hieu", customer["K5"].value)
        self.assertIn("Trung_bay", customer["L5"].value)
        self.assertIn("Khong_dat", customer["M5"].value)
        self.assertIn("Chi_tiet_Anh_Checkin", customer["K5"].value)

        replace_summary_rows(summary, facts, customer_end)
        self.assertIn('"KHTC"', summary["D5"].value)
        self.assertIn('"KHĐĐK"', summary["D5"].value)
        self.assertIn("MIN", summary["M5"].value)
        self.assertIn("$B$10", summary["P5"].value)
        self.assertEqual(summary.freeze_panes, "A5")
        workbook.close()

    def test_empty_customer_frame_keeps_valid_table_contract(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Chi_tiet_Khach_hang"
        for column in range(1, max(CUSTOMER_HEADERS) + 1):
            sheet.cell(4, column, f"H{column}")
        end_row = replace_customer_rows(sheet, pd.DataFrame())
        self.assertEqual(end_row, 4)
        self.assertEqual(sheet.auto_filter.ref, "A4:U4")
        workbook.close()


if __name__ == "__main__":
    unittest.main()
