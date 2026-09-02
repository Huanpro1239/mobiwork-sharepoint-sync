from __future__ import annotations

import unittest

import pandas as pd
from openpyxl import Workbook

from kpi.kpi_exporter import KPIExporter
from kpi.manual_labels import ManualLabelIndex
from kpi.review_queue import partition_review_rows, summarize_review_rows


class KPIReviewQueueTests(unittest.TestCase):
    def setUp(self) -> None:
        self.frame = pd.DataFrame(
            [
                {
                    "record_id": "needs-review",
                    "ten_nhan_vien": "NV NEEDS REVIEW",
                    "Phân Loại AI": "Can_duyet",
                    "Trạng Thái Quyết Định": "REVIEW_VALIDITY",
                    "hinh_anh": "https://example/review.jpg",
                },
                {
                    "record_id": "reviewed",
                    "ten_nhan_vien": "NV REVIEWED",
                    "Phân Loại AI": "Can_duyet",
                    "Trạng Thái Quyết Định": "REVIEW_FRAUD",
                    "hinh_anh": "https://example/reviewed.jpg",
                },
                {
                    "record_id": "pending",
                    "ten_nhan_vien": "NV PENDING",
                    "Phân Loại AI": "Khong_the_cham",
                    "Trạng Thái Quyết Định": "PENDING_SCORE",
                    "hinh_anh": "https://example/pending.jpg",
                },
                {
                    "record_id": "technical",
                    "ten_nhan_vien": "NV TECHNICAL",
                    "Phân Loại AI": "Khong_the_cham",
                    "Trạng Thái Quyết Định": "TECHNICAL_FAILURE",
                    "hinh_anh": "https://example/broken.jpg",
                },
                {
                    "record_id": "pass",
                    "ten_nhan_vien": "NV PASS",
                    "Phân Loại AI": "Bien_hieu",
                    "Trạng Thái Quyết Định": "AUTO_PASS_HIGH_CONFIDENCE",
                    "hinh_anh": "https://example/pass.jpg",
                },
                {
                    "record_id": "fail",
                    "ten_nhan_vien": "NV FAIL",
                    "Phân Loại AI": "Khong_dat",
                    "Trạng Thái Quyết Định": "AUTO_FAIL_LOW_EVIDENCE",
                    "hinh_anh": "https://example/fail.jpg",
                },
            ]
        )
        self.labels = ManualLabelIndex(
            by_record_id={"reviewed": "Khong_dat"},
            by_exact_key={},
            by_fallback_key={},
        )

    def test_partitions_manual_pending_and_technical_states_independently(self):
        partitions = partition_review_rows(self.frame, self.labels)

        self.assertEqual(partitions.manual_required["record_id"].tolist(), ["needs-review"])
        self.assertEqual(partitions.manual_resolved["record_id"].tolist(), ["reviewed"])
        self.assertEqual(partitions.pending["record_id"].tolist(), ["pending"])
        self.assertEqual(partitions.technical["record_id"].tolist(), ["technical"])
        self.assertTrue(partitions.fraud_audit.empty)
        self.assertTrue(partitions.historical_review.empty)

    def test_manual_label_excludes_review_row_from_required_queue(self):
        partitions = partition_review_rows(self.frame, self.labels)

        self.assertNotIn("reviewed", partitions.manual_required["record_id"].tolist())
        self.assertEqual(
            partitions.manual_resolved.iloc[0]["_manual_label"],
            "Khong_dat",
        )

    def test_summary_excludes_pending_and_technical_from_scored_denominator(self):
        summary = summarize_review_rows(self.frame, self.labels)

        self.assertEqual(summary["scored_decision_count"], 4)
        self.assertEqual(summary["manual_review_decision_count"], 2)
        self.assertEqual(summary["manual_review_required_count"], 1)
        self.assertEqual(summary["manual_review_resolved_count"], 1)
        self.assertEqual(summary["pending_score_count"], 1)
        self.assertEqual(summary["technical_failure_count"], 1)
        self.assertEqual(summary["auto_pass_count"], 1)
        self.assertEqual(summary["auto_fail_count"], 1)
        self.assertAlmostEqual(summary["manual_review_rate"], 0.5)
        self.assertAlmostEqual(summary["auto_pass_rate"], 0.25)

    def test_period_scope_removes_historical_and_redundant_review_from_current_kpi_queue(self):
        frame = pd.DataFrame(
            [
                {
                    "record_id": "historical-review",
                    "ngay": "2026-08-31",
                    "ma_kh": "OLD",
                    "Loại Cảnh": "Bien_hieu",
                    "Phân Loại AI": "Can_duyet",
                    "Trạng Thái Quyết Định": "TIER4_WEIGHTED_REVIEW",
                    "hinh_anh": "https://example/historical.jpg",
                },
                {
                    "record_id": "current-review",
                    "ngay": "2026-09-02",
                    "ma_kh": "CUR",
                    "Loại Cảnh": "Trung_bay",
                    "Phân Loại AI": "Can_duyet",
                    "Trạng Thái Quyết Định": "TIER4_WEIGHTED_REVIEW",
                    "hinh_anh": "https://example/current.jpg",
                },
                {
                    "record_id": "fraud-review",
                    "ngay": "2026-09-02",
                    "ma_kh": "FRAUD",
                    "Loại Cảnh": "Trung_bay",
                    "Phân Loại AI": "Can_duyet",
                    "Trạng Thái Quyết Định": "TIER0_REVIEW_FRAUD",
                    "hinh_anh": "https://example/fraud.jpg",
                },
                {
                    "record_id": "scene-pass",
                    "ngay": "2026-09-02",
                    "ma_kh": "REDUNDANT",
                    "Loại Cảnh": "Bien_hieu",
                    "Phân Loại AI": "Bien_hieu",
                    "Trạng Thái Quyết Định": "TIER1_HIGH_PASS",
                    "hinh_anh": "https://example/pass.jpg",
                },
                {
                    "record_id": "redundant-review",
                    "ngay": "2026-09-02",
                    "ma_kh": "REDUNDANT",
                    "Loại Cảnh": "Bien_hieu",
                    "Phân Loại AI": "Can_duyet",
                    "Trạng Thái Quyết Định": "REVIEW_NOVELTY",
                    "hinh_anh": "https://example/redundant.jpg",
                },
            ]
        )
        empty = ManualLabelIndex.empty()
        period = pd.Timestamp("2026-09-01")
        partitions = partition_review_rows(frame, empty, period_start=period)
        summary = summarize_review_rows(frame, empty, period_start=period)

        self.assertEqual(partitions.manual_required["record_id"].tolist(), ["current-review"])
        self.assertEqual(partitions.fraud_audit["record_id"].tolist(), ["fraud-review"])
        self.assertEqual(partitions.deferred_review["record_id"].tolist(), ["redundant-review"])
        self.assertEqual(partitions.historical_review["record_id"].tolist(), ["historical-review"])
        self.assertEqual(summary["manual_review_decision_count"], 4)
        self.assertEqual(summary["manual_review_required_count"], 1)
        self.assertEqual(summary["fraud_audit_required_count"], 1)
        self.assertEqual(summary["deferred_review_count"], 1)
        self.assertEqual(summary["historical_review_count"], 1)
        self.assertEqual(summary["current_period_scored_count"], 4)
        self.assertAlmostEqual(summary["current_period_operational_review_rate"], 0.5)

    def test_tiered_statuses_feed_metrics_and_manual_review_queue(self):
        tiered = pd.DataFrame(
            [
                {
                    "record_id": "tier-pass-1",
                    "Phân Loại AI": "Bien_hieu",
                    "Trạng Thái Quyết Định": "TIER1_HIGH_PASS",
                    "hinh_anh": "https://example/tier-pass-1.jpg",
                },
                {
                    "record_id": "tier-pass-2",
                    "Phân Loại AI": "Trung_bay",
                    "Trạng Thái Quyết Định": "TIER2_CONSENSUS_PASS",
                    "hinh_anh": "https://example/tier-pass-2.jpg",
                },
                {
                    "record_id": "tier-fail-1",
                    "Phân Loại AI": "Khong_dat",
                    "Trạng Thái Quyết Định": "TIER0_AUTO_FAIL_FRAUD",
                    "hinh_anh": "https://example/tier-fail-1.jpg",
                },
                {
                    "record_id": "tier-fail-2",
                    "Phân Loại AI": "Khong_dat",
                    "Trạng Thái Quyết Định": "TIER4_WEIGHTED_FAIL",
                    "hinh_anh": "https://example/tier-fail-2.jpg",
                },
                {
                    "record_id": "tier-review-fraud",
                    "Phân Loại AI": "Can_duyet",
                    "Trạng Thái Quyết Định": "TIER0_REVIEW_FRAUD",
                    "hinh_anh": "https://example/tier-review-fraud.jpg",
                },
                {
                    "record_id": "tier-review-weighted",
                    "Phân Loại AI": "Can_duyet",
                    "Trạng Thái Quyết Định": "TIER4_WEIGHTED_REVIEW",
                    "hinh_anh": "https://example/tier-review-weighted.jpg",
                },
            ]
        )

        summary = summarize_review_rows(tiered, ManualLabelIndex.empty())
        partitions = partition_review_rows(tiered, ManualLabelIndex.empty())

        self.assertEqual(summary["scored_decision_count"], 6)
        self.assertEqual(summary["auto_pass_count"], 2)
        self.assertEqual(summary["auto_fail_count"], 2)
        self.assertEqual(summary["manual_review_decision_count"], 2)
        self.assertAlmostEqual(summary["auto_pass_rate"], 2 / 6)
        self.assertAlmostEqual(summary["manual_review_rate"], 2 / 6)
        self.assertEqual(
            partitions.manual_required["record_id"].tolist(),
            ["tier-review-fraud", "tier-review-weighted"],
        )

    def test_operational_unique_counts_prefer_url_over_shared_image_bytes(self):
        technical = pd.DataFrame(
            [
                {
                    "record_id": "broken-a",
                    "hinh_anh": "https://example/a.jpg",
                    "image_sha256": "same-bytes",
                    "Phân Loại AI": "Khong_the_cham",
                    "Trạng Thái Quyết Định": "TECHNICAL_FAILURE",
                },
                {
                    "record_id": "broken-b",
                    "hinh_anh": "https://example/b.jpg",
                    "image_sha256": "same-bytes",
                    "Phân Loại AI": "Khong_the_cham",
                    "Trạng Thái Quyết Định": "TECHNICAL_FAILURE",
                },
            ]
        )

        summary = summarize_review_rows(technical, ManualLabelIndex.empty())

        self.assertEqual(summary["technical_failure_unique"], 2)

    def test_alert_sheet_renders_six_operational_sections(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Canh_bao"
        sheet.cell(1, 1, "2. DANH SÁCH ẢNH CẦN XỬ LÝ")
        sheet.cell(2, 1, "old header")
        sheet.cell(3, 1, "old data")
        sheet.cell(4, 1, "3. CẢNH BÁO KHÁC")

        KPIExporter._update_review_alerts(sheet, self.frame, self.labels)

        headings = {
            str(sheet.cell(row, 1).value): row
            for row in range(1, sheet.max_row + 1)
            if sheet.cell(row, 1).value
        }
        review_heading = next(key for key in headings if key.startswith("2A."))
        fraud_heading = next(key for key in headings if key.startswith("2B."))
        deferred_heading = next(key for key in headings if key.startswith("2C."))
        history_heading = next(key for key in headings if key.startswith("2D."))
        technical_heading = next(key for key in headings if key.startswith("2E."))
        pending_heading = next(key for key in headings if key.startswith("2F."))
        section_three = next(key for key in headings if key.startswith("3."))

        def employee_names(start_heading: str, end_heading: str) -> list[str]:
            return [
                str(sheet.cell(row, 2).value or "")
                for row in range(headings[start_heading] + 1, headings[end_heading])
            ]

        review_names = employee_names(review_heading, fraud_heading)
        technical_names = employee_names(technical_heading, pending_heading)
        pending_names = employee_names(pending_heading, section_three)

        self.assertIn("NV NEEDS REVIEW", review_names)
        self.assertNotIn("NV REVIEWED", review_names)
        self.assertNotIn("NV PENDING", review_names)
        self.assertNotIn("NV TECHNICAL", review_names)
        self.assertIn("NV TECHNICAL", technical_names)
        self.assertIn("NV PENDING", pending_names)
        self.assertLess(headings[fraud_heading], headings[deferred_heading])
        self.assertLess(headings[deferred_heading], headings[history_heading])
        workbook.close()

    def test_rescore_changes_ai_column_but_preserves_manual_override_formula(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Chi_tiet_Anh_Checkin"
        for column in range(1, 34):
            sheet.cell(4, column, f"H{column}")
            sheet.cell(5, column, "")
        exporter = object.__new__(KPIExporter)
        rescored = pd.DataFrame(
            [
                {
                    "record_id": "reviewed",
                    "Phân Loại AI": "Bien_hieu",
                    "Trạng Thái Quyết Định": "AUTO_PASS",
                    "hinh_anh": "https://example/reviewed.jpg",
                }
            ]
        )

        exporter._replace_detail_rows(sheet, rescored, self.labels, customer_end_row=4)

        self.assertEqual(sheet["G5"].value, "Bien_hieu")
        self.assertEqual(sheet["H5"].value, "Khong_dat")
        self.assertEqual(sheet["I5"].value, '=IF(H5<>"",H5,G5)')
        self.assertEqual(sheet["AE4"].value, "Nhóm Xử Lý")
        self.assertEqual(sheet["AG4"].value, "Hướng Dẫn Xử Lý")
        workbook.close()


if __name__ == "__main__":
    unittest.main()
