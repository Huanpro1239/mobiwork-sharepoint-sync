from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

import openpyxl

from src.planning.excel_io import write_shadow_workbook


class PlanningExcelIOTests(unittest.TestCase):
    def test_write_shadow_workbook_preserves_and_sorts_date_keys(self):
        rows = [
            {
                "Ma SP": "P1",
                "2026-08-29": 100,
            },
            {
                "Ma SP": "P2",
                "2026-08-01": 200,
                "SL chua xep": 0,
            },
        ]

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "shadow.xlsx"
            write_shadow_workbook({"Schedule": rows}, path)

            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb["Schedule"]
            headers = [cell.value for cell in ws[1]]

            self.assertEqual(
                headers,
                ["Ma SP", "SL chua xep", "2026-08-01", "2026-08-29"],
            )
            self.assertEqual(ws.cell(2, 4).value, 100)
            self.assertIsNone(ws.cell(2, 3).value)
            self.assertEqual(ws.cell(3, 2).value, 0)
            self.assertEqual(ws.cell(3, 3).value, 200)
            wb.close()

    def test_write_shadow_workbook_cleans_float_noise(self):
        rows = [
            {"Ma SP": "P1", "Qty": 26000.000000000004, "Ratio": 0.07999999999992724},
        ]

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "shadow.xlsx"
            write_shadow_workbook({"Data": rows}, path)

            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb["Data"]
            self.assertEqual(ws["B2"].value, 26000)
            self.assertEqual(ws["C2"].value, 0.08)
            self.assertEqual(ws["B2"].number_format, "#,##0.###")
            wb.close()


if __name__ == "__main__":
    unittest.main()
