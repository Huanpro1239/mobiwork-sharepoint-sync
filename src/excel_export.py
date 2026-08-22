from __future__ import annotations

import math
from datetime import date
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


CODE_COLUMNS = {
    "ma_phieu",
    "ma_kh",
    "sdt",
    "ma_nhom",
    "ma_nv_dat",
    "ma_nv_duyet",
    "ma_sp",
    "ma_kho_xuat",
}
DATE_COLUMNS = {"ngay_dat", "ngay_duyet", "ngay_tao", "ngay_sua"}
HEADER_NUMERIC_COLUMNS = {
    "tong_tien_hang",
    "tong_tien_vat",
    "tong_ck_sp",
    "ck_don_hang",
    "phai_thanh_toan",
}
DETAIL_NUMERIC_COLUMNS = {
    "stt",
    "so_luong",
    "don_gia",
    "thanh_tien",
    "chiet_khau",
    "vat",
}


def _normalize_codes(frame: pd.DataFrame) -> pd.DataFrame:
    for column in CODE_COLUMNS.intersection(frame.columns):
        frame[column] = frame[column].astype("string")
    return frame


def _normalize_numeric(frame: pd.DataFrame, columns: set[str]) -> pd.DataFrame:
    for column in columns.intersection(frame.columns):
        frame[column] = pd.to_numeric(frame[column], errors="coerce")
    return frame


def _normalize_dates(frame: pd.DataFrame) -> pd.DataFrame:
    for column in DATE_COLUMNS.intersection(frame.columns):
        parsed = pd.to_datetime(frame[column], errors="coerce", utc=True)
        frame[column] = parsed.dt.tz_convert("Asia/Ho_Chi_Minh").dt.tz_localize(None)
    return frame


def _assert_unique(frame: pd.DataFrame, columns: list[str], label: str) -> None:
    if frame.empty or any(column not in frame.columns for column in columns):
        return
    key_frame = frame[columns]
    missing_mask = key_frame.isna().any(axis=1) | key_frame.astype("string").eq("").any(axis=1)
    if missing_mask.any():
        rows = [str(index + 2) for index in frame.index[missing_mask][:10]]
        raise ValueError(f"{label}: empty key {columns} at Excel rows {', '.join(rows)}")
    duplicate_mask = frame.duplicated(subset=columns, keep=False)
    if duplicate_mask.any():
        sample = frame.loc[duplicate_mask, columns].head(5).to_dict("records")
        raise ValueError(f"{label}: duplicate key {columns}; sample={sample}")


def _validate_excel_size(frame: pd.DataFrame, label: str) -> None:
    # One row is reserved for the header.
    if len(frame) > 1_048_575:
        raise ValueError(
            f"{label} contains {len(frame):,} rows and exceeds the Excel worksheet limit"
        )


def _valid_line_number(value: Any) -> int | None:
    """Return a positive integer line number or None when MobiWork did not provide one."""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, str) and not value.strip():
        return None
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(numeric) or numeric <= 0 or not numeric.is_integer():
        return None
    return int(numeric)


def _normalize_order_items(order: dict[str, Any]) -> list[dict[str, Any]]:
    """Normalize sold/promo lines and fill only missing/invalid ``stt`` values.

    Some historical MobiWork Order/Bill rows omit ``stt`` inside ``san_pham``.
    A line without ``stt`` must not be dropped, but ``ma_phieu + stt`` still needs to
    remain a deterministic business key. Existing valid line numbers are preserved;
    missing/invalid values receive the first unused positive integer in source order.
    Duplicate line numbers supplied by MobiWork are intentionally not repaired and are
    still rejected by the downstream uniqueness check.
    """
    standard_items = order.get("san_pham") or []
    promo_items = order.get("san_pham_km") or []
    if not isinstance(standard_items, list):
        raise TypeError("san_pham must be a list")
    if not isinstance(promo_items, list):
        raise TypeError("san_pham_km must be a list when present")

    normalized: list[dict[str, Any]] = []
    for item in standard_items:
        if not isinstance(item, dict):
            raise TypeError("san_pham contains a non-object item")
        normalized.append(dict(item))

    for item in promo_items:
        if not isinstance(item, dict):
            raise TypeError("san_pham_km contains a non-object item")
        promo_item = dict(item)
        promo_item.setdefault("is_km", True)
        normalized.append(promo_item)

    used = {
        line_number
        for item in normalized
        if (line_number := _valid_line_number(item.get("stt"))) is not None
    }
    candidate = 1
    for item in normalized:
        line_number = _valid_line_number(item.get("stt"))
        if line_number is not None:
            item["stt"] = line_number
            continue
        while candidate in used:
            candidate += 1
        item["stt"] = candidate
        used.add(candidate)
        candidate += 1

    return normalized


def build_order_frames(
    records: list[dict[str, Any]],
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Build normalized order header and line-item tables.

    MobiWork represents both sold and promotional products inside ``san_pham``.
    ``is_km`` remains on the line so downstream Excel/Power BI users can filter
    promotional items without splitting one business entity across two tables.
    ``san_pham_km`` is still accepted as a backwards-compatible source and is
    normalized into the same detail table.
    """
    header_rows: list[dict[str, Any]] = []
    detail_rows: list[dict[str, Any]] = []

    for order in records:
        header = {
            key: value
            for key, value in order.items()
            if key not in {"san_pham", "san_pham_km"}
        }
        header_rows.append(header)

        parent_ma_phieu = header.get("ma_phieu")
        for item in _normalize_order_items(order):
            detail = {**header, **item}
            # Some MobiWork child objects contain ma_phieu=null/"".  The document
            # number belongs to the parent order/bill, so an empty child value must
            # never erase the valid parent business key.
            child_ma_phieu = detail.get("ma_phieu")
            if child_ma_phieu is None or (
                isinstance(child_ma_phieu, str) and not child_ma_phieu.strip()
            ):
                detail["ma_phieu"] = parent_ma_phieu
            detail_rows.append(detail)

    header_frame = (
        pd.json_normalize(header_rows, sep="_") if header_rows else pd.DataFrame()
    )
    detail_frame = (
        pd.json_normalize(detail_rows, sep="_") if detail_rows else pd.DataFrame()
    )

    header_frame = _normalize_codes(header_frame)
    header_frame = _normalize_numeric(header_frame, HEADER_NUMERIC_COLUMNS)
    header_frame = _normalize_dates(header_frame)

    detail_frame = _normalize_codes(detail_frame)
    detail_frame = _normalize_numeric(detail_frame, DETAIL_NUMERIC_COLUMNS | HEADER_NUMERIC_COLUMNS)
    detail_frame = _normalize_dates(detail_frame)
    if "is_km" in detail_frame.columns:
        detail_frame["is_km"] = detail_frame["is_km"].astype("boolean")
        insert_at = detail_frame.columns.get_loc("is_km") + 1
        detail_frame.insert(
            insert_at,
            "loai_hang",
            detail_frame["is_km"].map({True: "Khuyến mãi", False: "Bán hàng"}),
        )

    _assert_unique(header_frame, ["ma_phieu"], "DonHang")
    _assert_unique(detail_frame, ["ma_phieu", "stt"], "ChiTietSP")
    _validate_excel_size(header_frame, "DonHang")
    _validate_excel_size(detail_frame, "ChiTietSP")
    return header_frame, detail_frame


def _format_sheet(writer: pd.ExcelWriter, sheet_name: str) -> None:
    worksheet = writer.book[sheet_name]
    worksheet.freeze_panes = "A2"
    if worksheet.max_row < 1 or worksheet.max_column < 1:
        return

    worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    header_map = {
        worksheet.cell(row=1, column=column).value: column
        for column in range(1, worksheet.max_column + 1)
    }

    for field_name in CODE_COLUMNS:
        column = header_map.get(field_name)
        if not column:
            continue
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row, column=column).number_format = "@"

    for field_name in DATE_COLUMNS:
        column = header_map.get(field_name)
        if not column:
            continue
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row, column=column).number_format = "yyyy-mm-dd hh:mm:ss"

    numeric_fields = HEADER_NUMERIC_COLUMNS | DETAIL_NUMERIC_COLUMNS
    for field_name in numeric_fields:
        column = header_map.get(field_name)
        if not column:
            continue
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row, column=column).number_format = "#,##0.00"

    # Keep files readable without producing extreme widths from long notes/addresses.
    sample_rows = min(worksheet.max_row, 200)
    for column in range(1, worksheet.max_column + 1):
        max_length = 0
        for row in range(1, sample_rows + 1):
            value = worksheet.cell(row=row, column=column).value
            if value is not None:
                max_length = max(max_length, len(str(value)))
        worksheet.column_dimensions[get_column_letter(column)].width = min(max(max_length + 2, 10), 40)


def export_excel(
    records: list[dict[str, Any]],
    report_name: str,
    target_date: date,
    export_mode: str = "flat",
    file_suffix: str | None = None,
) -> Path:
    output_dir = Path("output")
    output_dir.mkdir(parents=True, exist_ok=True)
    suffix = file_suffix or f"{target_date:%Y-%m-%d}"
    path = output_dir / f"{report_name}_{suffix}.xlsx"

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        if export_mode == "order":
            header, detail = build_order_frames(records)
            header.to_excel(writer, sheet_name="DonHang", index=False)
            detail.to_excel(writer, sheet_name="ChiTietSP", index=False)
            for sheet_name in ("DonHang", "ChiTietSP"):
                _format_sheet(writer, sheet_name)
        else:
            frame = pd.json_normalize(records, sep="_") if records else pd.DataFrame()
            _validate_excel_size(frame, "Data")
            frame.to_excel(writer, sheet_name="Data", index=False)
            _format_sheet(writer, "Data")

    return path
