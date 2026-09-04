from __future__ import annotations

import re
from datetime import date, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.utils import get_column_letter

from excel_export import _assert_unique, _format_sheet, _validate_excel_size, build_order_frames


SYNC_DATE_COLUMN = "_sync_date"


def master_filename(report_name: str, target_date: date) -> str:
    return f"{report_name}_{target_date:%Y-%m}.xlsx"


def month_dates_through(target_date: date) -> list[date]:
    first = target_date.replace(day=1)
    return [first + timedelta(days=offset) for offset in range((target_date - first).days + 1)]


def _empty_frames(export_mode: str) -> dict[str, pd.DataFrame]:
    if export_mode == "order":
        return {"DonHang": pd.DataFrame(), "ChiTietSP": pd.DataFrame()}
    return {"Data": pd.DataFrame()}


def frames_from_records(
    records: list[dict[str, Any]],
    export_mode: str,
    target_date: date,
) -> dict[str, pd.DataFrame]:
    partition = target_date.isoformat()
    if export_mode == "order":
        header, detail = build_order_frames(records)
        header.insert(0, SYNC_DATE_COLUMN, partition)
        detail.insert(0, SYNC_DATE_COLUMN, partition)
        return {"DonHang": header, "ChiTietSP": detail}

    frame = pd.json_normalize(records, sep="_") if records else pd.DataFrame()
    frame.insert(0, SYNC_DATE_COLUMN, partition)
    _validate_excel_size(frame, "Data")
    return {"Data": frame}


def read_master(content: bytes, export_mode: str) -> dict[str, pd.DataFrame]:
    if not content:
        return _empty_frames(export_mode)

    source = BytesIO(content)
    if export_mode == "order":
        frames = {
            "DonHang": pd.read_excel(source, sheet_name="DonHang", engine="openpyxl"),
        }
        source.seek(0)
        frames["ChiTietSP"] = pd.read_excel(
            source,
            sheet_name="ChiTietSP",
            engine="openpyxl",
        )
    else:
        frames = {"Data": pd.read_excel(source, sheet_name="Data", engine="openpyxl")}

    for sheet_name, frame in frames.items():
        if SYNC_DATE_COLUMN not in frame.columns:
            raise ValueError(
                f"Monthly master sheet {sheet_name!r} is missing {SYNC_DATE_COLUMN}; "
                "a one-time month rebuild is required"
            )
        frame[SYNC_DATE_COLUMN] = frame[SYNC_DATE_COLUMN].astype("string")
    return frames


def _combine_partition_frames(old: pd.DataFrame, new: pd.DataFrame) -> pd.DataFrame:
    """Combine two partitions without passing empty frames through pandas concat."""
    if old.empty:
        return new.reset_index(drop=True).copy()
    if new.empty:
        return old.reset_index(drop=True).copy()
    return pd.concat([old, new], ignore_index=True, sort=False)


def _normalize_key_value(value: Any) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def _incoming_business_keys(
    frame: pd.DataFrame,
    keys: list[str],
    label: str,
) -> set[tuple[str, ...]]:
    if frame.empty or not keys:
        return set()
    missing = [key for key in keys if key not in frame.columns]
    if missing:
        raise ValueError(f"{label} is missing configured upsert key(s): {missing}")

    values: set[tuple[str, ...]] = set()
    for row_number, row in enumerate(
        frame.loc[:, keys].itertuples(index=False, name=None),
        start=1,
    ):
        normalized = tuple(_normalize_key_value(value) for value in row)
        if any(not value for value in normalized):
            raise ValueError(
                f"{label} row {row_number} has an empty configured upsert key {keys}"
            )
        values.add(normalized)
    return values


def _existing_business_key(row: pd.Series, keys: list[str]) -> tuple[str, ...] | None:
    values = tuple(_normalize_key_value(row.get(key)) for key in keys)
    return None if any(not value for value in values) else values


def _partition_rows(frame: pd.DataFrame, partition: str, label: str) -> pd.DataFrame:
    if frame.empty:
        return frame.copy()
    if SYNC_DATE_COLUMN not in frame.columns:
        raise ValueError(f"{label} is missing {SYNC_DATE_COLUMN}")
    return frame.loc[frame[SYNC_DATE_COLUMN].astype("string") == partition].copy()


def _assert_partition_applied(
    merged: dict[str, pd.DataFrame],
    incoming: dict[str, pd.DataFrame],
    target_date: date,
    export_mode: str,
    configured_keys: list[str],
) -> None:
    """Fail before publish if a fetched partition was lost or only partly applied.

    This is intentionally checked after every merge, not only during full-month rebuilds.
    It protects hourly, nightly, weekly and month-close paths with the same invariant:
    everything fetched for the target partition must be represented in the resulting
    monthly master, while stale rows from an older version of that partition must be gone.
    """
    partition = target_date.isoformat()

    if export_mode == "order":
        incoming_header = incoming.get("DonHang", pd.DataFrame())
        incoming_detail = incoming.get("ChiTietSP", pd.DataFrame())
        merged_header = merged.get("DonHang", pd.DataFrame())
        merged_detail = merged.get("ChiTietSP", pd.DataFrame())

        expected_header_keys = _incoming_business_keys(
            incoming_header,
            ["ma_phieu"],
            "DonHang incoming quality gate",
        )
        expected_detail_keys = _incoming_business_keys(
            incoming_detail,
            ["ma_phieu", "stt"],
            "ChiTietSP incoming quality gate",
        )

        applied_header = _partition_rows(
            merged_header,
            partition,
            "DonHang merged quality gate",
        )
        applied_detail = _partition_rows(
            merged_detail,
            partition,
            "ChiTietSP merged quality gate",
        )
        applied_header_keys = _incoming_business_keys(
            applied_header,
            ["ma_phieu"],
            "DonHang applied quality gate",
        )
        applied_detail_keys = _incoming_business_keys(
            applied_detail,
            ["ma_phieu", "stt"],
            "ChiTietSP applied quality gate",
        )

        if applied_header_keys != expected_header_keys:
            missing = sorted(expected_header_keys - applied_header_keys)
            unexpected = sorted(applied_header_keys - expected_header_keys)
            raise RuntimeError(
                "Monthly master quality gate failed for DonHang partition "
                f"{partition}: missing={missing[:10]} unexpected={unexpected[:10]}"
            )
        if applied_detail_keys != expected_detail_keys:
            missing = sorted(expected_detail_keys - applied_detail_keys)
            unexpected = sorted(applied_detail_keys - expected_detail_keys)
            raise RuntimeError(
                "Monthly master quality gate failed for ChiTietSP partition "
                f"{partition}: missing={missing[:10]} unexpected={unexpected[:10]}"
            )

        # A repeated ma_phieu on a later source date is a full replacement of that
        # business entity. No old detail line for the same document may survive on
        # another _sync_date after the new partition is applied.
        if expected_header_keys and not merged_detail.empty:
            if "ma_phieu" not in merged_detail.columns:
                raise RuntimeError(
                    "Monthly master quality gate failed: ChiTietSP has no ma_phieu"
                )
            parent_values = {key[0] for key in expected_header_keys}
            normalized_parent = merged_detail["ma_phieu"].map(_normalize_key_value)
            all_current_detail = merged_detail.loc[normalized_parent.isin(parent_values)]
            all_current_detail_keys = _incoming_business_keys(
                all_current_detail,
                ["ma_phieu", "stt"],
                "ChiTietSP current-business quality gate",
            )
            if all_current_detail_keys != expected_detail_keys:
                stale = sorted(all_current_detail_keys - expected_detail_keys)
                missing = sorted(expected_detail_keys - all_current_detail_keys)
                raise RuntimeError(
                    "Monthly master quality gate failed for replaced order detail "
                    f"partition {partition}: missing={missing[:10]} stale={stale[:10]}"
                )
        return

    incoming_data = incoming.get("Data", pd.DataFrame())
    merged_data = merged.get("Data", pd.DataFrame())
    applied_data = _partition_rows(merged_data, partition, "Data merged quality gate")

    if configured_keys:
        expected_keys = _incoming_business_keys(
            incoming_data,
            configured_keys,
            "Data incoming quality gate",
        )
        applied_keys = _incoming_business_keys(
            applied_data,
            configured_keys,
            "Data applied quality gate",
        )
        if applied_keys != expected_keys:
            missing = sorted(expected_keys - applied_keys)
            unexpected = sorted(applied_keys - expected_keys)
            raise RuntimeError(
                "Monthly master quality gate failed for Data partition "
                f"{partition}: missing={missing[:10]} unexpected={unexpected[:10]}"
            )
        return

    if len(applied_data) != len(incoming_data):
        raise RuntimeError(
            "Monthly master quality gate failed for Data partition "
            f"{partition}: fetched_rows={len(incoming_data)} stored_rows={len(applied_data)}"
        )


def merge_partition(
    existing: dict[str, pd.DataFrame],
    incoming: dict[str, pd.DataFrame],
    target_date: date,
    export_mode: str,
    upsert_keys: list[str] | None = None,
) -> dict[str, pd.DataFrame]:
    partition = target_date.isoformat()
    sheet_names = ("DonHang", "ChiTietSP") if export_mode == "order" else ("Data",)
    merged: dict[str, pd.DataFrame] = {}

    configured_keys = list(upsert_keys or ([] if export_mode == "flat" else ["ma_phieu"]))
    if export_mode == "order" and configured_keys not in ([], ["ma_phieu"]):
        raise ValueError(
            "Order-mode monthly masters currently require upsert_keys=['ma_phieu']"
        )

    source_sheet = "DonHang" if export_mode == "order" else "Data"
    incoming_keys = _incoming_business_keys(
        incoming.get(source_sheet, pd.DataFrame()),
        configured_keys,
        f"{source_sheet} incoming partition",
    )

    for sheet_name in sheet_names:
        old = existing.get(sheet_name, pd.DataFrame()).copy()
        new = incoming.get(sheet_name, pd.DataFrame()).copy()
        if not old.empty:
            if SYNC_DATE_COLUMN not in old.columns:
                raise ValueError(
                    f"Monthly master sheet {sheet_name!r} is missing {SYNC_DATE_COLUMN}"
                )
            mask = old[SYNC_DATE_COLUMN].astype("string") != partition
            if incoming_keys and all(key in old.columns for key in configured_keys):
                old_keys = old.apply(
                    lambda row: _existing_business_key(row, configured_keys),
                    axis=1,
                )
                mask = mask & (~old_keys.isin(incoming_keys))
            old = old.loc[mask]
        combined = _combine_partition_frames(old, new)
        if SYNC_DATE_COLUMN in combined.columns:
            combined[SYNC_DATE_COLUMN] = combined[SYNC_DATE_COLUMN].astype("string")
            combined = combined.sort_values(SYNC_DATE_COLUMN, kind="stable").reset_index(drop=True)
        _validate_excel_size(combined, sheet_name)
        merged[sheet_name] = combined

    if export_mode == "order":
        _assert_unique(merged["DonHang"], ["ma_phieu"], "DonHang monthly master")
        _assert_unique(
            merged["ChiTietSP"],
            ["ma_phieu", "stt"],
            "ChiTietSP monthly master",
        )
    elif configured_keys:
        _assert_unique(merged["Data"], configured_keys, "Data monthly master")

    _assert_partition_applied(
        merged,
        incoming,
        target_date,
        export_mode,
        configured_keys,
    )
    return merged


def build_month_from_partitions(
    partitions: list[tuple[date, list[dict[str, Any]]]],
    export_mode: str,
    upsert_keys: list[str] | None = None,
) -> dict[str, pd.DataFrame]:
    master = _empty_frames(export_mode)
    for target_date, records in partitions:
        incoming = frames_from_records(records, export_mode, target_date)
        master = merge_partition(
            master,
            incoming,
            target_date,
            export_mode,
            upsert_keys=upsert_keys,
        )
    return master


def write_master(
    frames: dict[str, pd.DataFrame],
    report_name: str,
    target_date: date,
) -> Path:
    output_dir = Path("output")
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / master_filename(report_name, target_date)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, frame in frames.items():
            frame.to_excel(writer, sheet_name=sheet_name, index=False)
            _format_sheet(writer, sheet_name)
            worksheet = writer.book[sheet_name]
            headers = {
                worksheet.cell(row=1, column=column).value: column
                for column in range(1, worksheet.max_column + 1)
            }
            sync_column = headers.get(SYNC_DATE_COLUMN)
            if sync_column:
                worksheet.column_dimensions[get_column_letter(sync_column)].hidden = True
    return path


def master_row_count(frames: dict[str, pd.DataFrame], export_mode: str) -> int:
    if export_mode == "order":
        return len(frames.get("DonHang", pd.DataFrame()))
    return len(frames.get("Data", pd.DataFrame()))


def is_legacy_report_file(name: str, report_name: str, canonical_name: str) -> bool:
    if name == canonical_name:
        return False
    if name.startswith(("__sync_tmp_", "__sync_backup_", "__sync_failed_")):
        return True
    escaped = re.escape(report_name)
    if re.fullmatch(rf"{escaped}_\d{{4}}-\d{{2}}-\d{{2}}\.xlsx", name):
        return True
    return bool(re.fullmatch(rf"{escaped}_History_.*\.xlsx", name, flags=re.IGNORECASE))
