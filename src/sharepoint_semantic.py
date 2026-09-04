from __future__ import annotations

import hashlib
import json
import logging
import math
import time
from datetime import date, datetime, time as dt_time
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from sharepoint import SharePointClient


LOG = logging.getLogger("mobiwork_sync")


def _normalize_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, bool)):
        return value
    if isinstance(value, float):
        if math.isnan(value):
            return {"type": "float", "value": "NaN"}
        if math.isinf(value):
            return {"type": "float", "value": "Infinity" if value > 0 else "-Infinity"}
        return {"type": "float", "value": format(value, ".17g")}
    if isinstance(value, (datetime, date, dt_time)):
        return {"type": type(value).__name__, "value": value.isoformat()}
    return {"type": type(value).__name__, "value": str(value)}


def workbook_semantic_fingerprint(content: bytes) -> tuple[str, list[dict[str, Any]]]:
    """Hash workbook business content while ignoring OOXML package metadata.

    SharePoint/Office can rewrite an .xlsx ZIP package after upload (document metadata,
    relationship parts, timestamps, etc.). Byte equality is therefore too strict for
    Excel files. This fingerprint covers sheet order/names and every non-empty cell's
    coordinate, data type and value, which is the business payload we need to protect.
    """
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=False)
    digest = hashlib.sha256()
    sheets: list[dict[str, Any]] = []

    try:
        for worksheet in workbook.worksheets:
            sheet_cells = 0
            max_data_row = 0
            max_data_column = 0
            digest.update(b"sheet\0")
            digest.update(worksheet.title.encode("utf-8"))
            digest.update(b"\0")

            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    sheet_cells += 1
                    max_data_row = max(max_data_row, cell.row)
                    max_data_column = max(max_data_column, cell.column)
                    record = [
                        worksheet.title,
                        cell.coordinate,
                        cell.data_type,
                        _normalize_value(cell.value),
                    ]
                    digest.update(
                        json.dumps(
                            record,
                            ensure_ascii=False,
                            separators=(",", ":"),
                            sort_keys=True,
                        ).encode("utf-8")
                    )
                    digest.update(b"\n")

            sheets.append(
                {
                    "name": worksheet.title,
                    "non_empty_cells": sheet_cells,
                    "data_rows": max_data_row,
                    "data_columns": max_data_column,
                }
            )
    finally:
        workbook.close()

    return digest.hexdigest(), sheets


def workbooks_semantically_equal(
    expected_content: bytes,
    actual_content: bytes,
) -> tuple[bool, dict[str, Any]]:
    expected_hash, expected_sheets = workbook_semantic_fingerprint(expected_content)
    actual_hash, actual_sheets = workbook_semantic_fingerprint(actual_content)
    details = {
        "expected_semantic_sha256": expected_hash,
        "actual_semantic_sha256": actual_hash,
        "expected_sheets": expected_sheets,
        "actual_sheets": actual_sheets,
    }
    return expected_hash == actual_hash and expected_sheets == actual_sheets, details


class SemanticSharePointClient(SharePointClient):
    """SharePoint client that verifies Excel business content, not ZIP bytes."""

    @staticmethod
    def _is_xlsx(filename: str) -> bool:
        return filename.casefold().endswith(".xlsx")

    def _put_content(
        self,
        drive_id: str,
        remote_folder: str,
        filename: str,
        content: bytes,
        content_type: str,
    ) -> dict[str, Any]:
        """Avoid staged replace when an existing Excel workbook is semantically unchanged."""
        if self._is_xlsx(filename):
            remote_path = "/".join(
                part for part in (remote_folder.strip("/"), filename) if part
            )
            existing = self.get_item_by_path(drive_id, remote_path)
            if existing:
                if "folder" in existing:
                    raise RuntimeError(
                        f"SharePoint target {remote_path!r} exists but is a folder"
                    )
                item_id = str(existing.get("id", "")).strip()
                if not item_id:
                    raise RuntimeError(
                        f"Existing SharePoint file {remote_path!r} has no driveItem id"
                    )

                current_content = self._download_item_content(drive_id, item_id)
                try:
                    matched, details = workbooks_semantically_equal(content, current_content)
                except Exception as exc:
                    LOG.warning(
                        "Unable to compare existing Excel workbook %s before upload: %s: %s. "
                        "Falling back to staged replacement.",
                        remote_path,
                        type(exc).__name__,
                        exc,
                    )
                else:
                    if matched:
                        LOG.info(
                            "Skipping unchanged SharePoint Excel write: %s "
                            "local_bytes=%s remote_bytes=%s",
                            remote_path,
                            len(content),
                            len(current_content),
                        )
                        return {
                            **existing,
                            **details,
                            "size": len(current_content),
                            "local_size": len(content),
                            "verification_mode": "xlsx_semantic_noop",
                            "semantic_match": True,
                            "upload_skipped": True,
                        }

        return super()._put_content(
            drive_id,
            remote_folder,
            filename,
            content,
            content_type,
        )

    def _verify_xlsx_content(
        self,
        drive_id: str,
        item_id: str,
        filename: str,
        expected_content: bytes,
        attempts: int = 3,
    ) -> tuple[bytes, dict[str, Any]]:
        last_details: dict[str, Any] = {}
        last_content = b""

        for attempt in range(attempts):
            last_content = self._download_item_content(drive_id, item_id)
            if last_content == expected_content:
                return last_content, {
                    "verification_mode": "byte_exact",
                    "semantic_match": True,
                }

            try:
                matched, last_details = workbooks_semantically_equal(
                    expected_content,
                    last_content,
                )
            except Exception as exc:
                last_details = {
                    "verification_error": f"{type(exc).__name__}: {exc}",
                }
                matched = False

            if matched:
                return last_content, {
                    **last_details,
                    "verification_mode": "xlsx_semantic",
                    "semantic_match": True,
                }

            if attempt < attempts - 1:
                delay = min(1.0 * (2**attempt), 3.0)
                LOG.warning(
                    "SharePoint Excel semantic verification not settled for %s. "
                    "Retry %s/%s in %.1fs",
                    filename,
                    attempt + 1,
                    attempts,
                    delay,
                )
                time.sleep(delay)

        raise RuntimeError(
            f"SharePoint Excel semantic mismatch for {filename}: "
            f"local_bytes={len(expected_content)}, remote_bytes={len(last_content)}, "
            f"details={json.dumps(last_details, ensure_ascii=False)[:1000]}"
        )

    def _verify_uploaded_size(
        self,
        drive_id: str,
        filename: str,
        uploaded: dict[str, Any],
        expected_size: int,
        verification_attempts: int = 3,
        expected_content: bytes | None = None,
    ) -> dict[str, Any]:
        if self._is_xlsx(filename) and expected_content is not None:
            item_id = str(uploaded.get("id", "")).strip()
            if not item_id:
                raise RuntimeError(
                    f"SharePoint Excel upload for {filename} returned no driveItem id"
                )

            actual_content, verification = self._verify_xlsx_content(
                drive_id,
                item_id,
                filename,
                expected_content,
                attempts=verification_attempts,
            )
            metadata = self._request("GET", self._item_url(drive_id, item_id)).json()
            result = {**uploaded, **metadata, **verification}
            result["size"] = len(actual_content)
            result["local_size"] = expected_size
            LOG.info(
                "Verified Excel workbook semantically: %s local_bytes=%s remote_bytes=%s mode=%s",
                filename,
                expected_size,
                len(actual_content),
                verification.get("verification_mode"),
            )
            return result

        return super()._verify_uploaded_size(
            drive_id,
            filename,
            uploaded,
            expected_size,
            verification_attempts=verification_attempts,
            expected_content=expected_content,
        )

    def _verify_exact_item_content(
        self,
        drive_id: str,
        item_id: str,
        filename: str,
        expected_content: bytes,
    ) -> dict[str, Any]:
        if not self._is_xlsx(filename):
            return super()._verify_exact_item_content(
                drive_id,
                item_id,
                filename,
                expected_content,
            )

        actual_content, verification = self._verify_xlsx_content(
            drive_id,
            item_id,
            filename,
            expected_content,
        )
        metadata = self._request("GET", self._item_url(drive_id, item_id)).json()
        result = {**metadata, **verification}
        result["size"] = len(actual_content)
        result["local_size"] = len(expected_content)
        return result
