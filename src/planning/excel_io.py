from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ISO_DATE_HEADER = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def column_letters(start: int, end: int) -> list[str]:
    return [get_column_letter(i) for i in range(start, end + 1)]


def read_sheet_rows(
    workbook_bytes: bytes,
    sheet_name: str,
    *,
    min_row: int,
    max_col: int,
    data_only: bool = True,
) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(
        BytesIO(workbook_bytes),
        read_only=True,
        data_only=data_only,
        keep_vba=True,
    )
    if sheet_name not in wb.sheetnames:
        raise KeyError(
            f"Sheet {sheet_name!r} not found. Available: {wb.sheetnames}"
        )
    ws = wb[sheet_name]
    letters = column_letters(1, max_col)
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(
        min_row=min_row,
        max_col=max_col,
        values_only=True,
    ):
        if not any(value not in (None, "") for value in values):
            continue
        rows.append({letters[i]: values[i] for i in range(len(values))})
    return rows


def read_table_by_header(
    workbook_bytes: bytes,
    sheet_name: str,
    *,
    header_row: int,
    min_row: int | None = None,
    data_only: bool = True,
) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(
        BytesIO(workbook_bytes),
        read_only=True,
        data_only=data_only,
        keep_vba=True,
    )
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[header_row]]
    last = max(
        (
            i
            for i, value in enumerate(headers, start=1)
            if value not in (None, "")
        ),
        default=0,
    )
    names = [
        str(value).strip() if value is not None else f"_col_{i}"
        for i, value in enumerate(headers[:last], start=1)
    ]
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(
        min_row=min_row or header_row + 1,
        max_col=last,
        values_only=True,
    ):
        if not any(value not in (None, "") for value in values):
            continue
        rows.append(dict(zip(names, values, strict=False)))
    return rows


def _ordered_union_headers(rows: list[dict[str, Any]]) -> list[str]:
    """Retain all keys while keeping schedule date columns chronological."""
    headers: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in row:
            if key in seen:
                continue
            seen.add(key)
            headers.append(key)

    date_headers = sorted(
        header for header in headers if ISO_DATE_HEADER.fullmatch(str(header))
    )
    static_headers = [
        header for header in headers if not ISO_DATE_HEADER.fullmatch(str(header))
    ]
    return static_headers + date_headers


def _clean_output_value(value: Any) -> Any:
    """Remove binary floating-point noise at the export boundary only."""
    if isinstance(value, float):
        if abs(value) < 1e-10:
            return 0
        nearest = round(value)
        if abs(value - nearest) < 1e-9:
            return int(nearest)
        return round(value, 6)
    return value


def _column_width(header: str, rows: list[dict[str, Any]]) -> float:
    if ISO_DATE_HEADER.fullmatch(header):
        return 12
    lowered = header.lower()
    if "ten " in lowered or lowered.startswith("ten") or "tên" in lowered:
        return 34
    if "ghi chu" in lowered or "ghi chú" in lowered or "de xuat" in lowered or "đề xuất" in lowered:
        return 38
    if "ma " in lowered or lowered.startswith("ma") or "mã" in lowered:
        return 16
    if "ngay" in lowered or "ngày" in lowered:
        return 15

    lengths = [len(header)]
    for row in rows[:100]:
        value = row.get(header)
        if value not in (None, ""):
            lengths.append(len(str(_clean_output_value(value))))
    return min(max(max(lengths) + 2, 11), 22)


def _format_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, headers: list[str], rows: list[dict[str, Any]]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E1F2")
    border = Border(bottom=thin)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 30

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for column_index, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(column_index)].width = _column_width(str(header), rows)
        is_date_column = bool(ISO_DATE_HEADER.fullmatch(str(header)))
        for cell in ws.iter_cols(
            min_col=column_index,
            max_col=column_index,
            min_row=2,
            max_row=ws.max_row,
        ):
            data_cell = cell[0]
            value = data_cell.value
            data_cell.alignment = Alignment(
                horizontal="center" if is_date_column else "left",
                vertical="center",
            )
            if isinstance(value, (date, datetime)):
                data_cell.number_format = "dd/mm/yyyy"
            elif isinstance(value, (int, float)) and not isinstance(value, bool):
                data_cell.number_format = "#,##0.###"
                if not is_date_column:
                    data_cell.alignment = Alignment(horizontal="right", vertical="center")


def write_shadow_workbook(
    tables: dict[str, list[dict[str, Any]]],
    output: Path,
) -> Path:
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for raw_name, rows in tables.items():
        name = raw_name[:31]
        ws = wb.create_sheet(name)
        if not rows:
            ws.append(["No data"])
            ws["A1"].font = Font(bold=True)
            continue
        headers = _ordered_union_headers(rows)
        ws.append(headers)
        for row in rows:
            ws.append([_clean_output_value(row.get(header)) for header in headers])
        _format_sheet(ws, headers, rows)
    wb.save(output)
    return output
