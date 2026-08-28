from __future__ import annotations

from io import BytesIO
from typing import Any, Iterable, Mapping

import openpyxl
from openpyxl.utils import get_column_letter

from .normalize import clean_text, normalize_code, normalize_compare_text, to_number
from .vba_port import aggregate_sales_actual, sales_channels_in_cases

Row = Mapping[str, Any]


def first_sheet_name(workbook_bytes: bytes) -> str:
    wb = openpyxl.load_workbook(
        BytesIO(workbook_bytes), read_only=True, data_only=True, keep_vba=True
    )
    return wb.sheetnames[0]


def sheet_name_by_index(workbook_bytes: bytes, index: int) -> str:
    """Return a 1-based worksheet name, matching VBA Worksheets(index)."""
    wb = openpyxl.load_workbook(
        BytesIO(workbook_bytes), read_only=True, data_only=True, keep_vba=True
    )
    if index < 1 or index > len(wb.sheetnames):
        raise IndexError(f"Worksheet index {index} out of range; count={len(wb.sheetnames)}")
    return wb.sheetnames[index - 1]


def find_column_by_header(
    workbook_bytes: bytes,
    sheet_name: str,
    header: str,
    *,
    first_row: int = 1,
    last_row: int = 9,
) -> str:
    """Match the VBA FindColumnByHeader behavior for source workbooks."""
    wb = openpyxl.load_workbook(
        BytesIO(workbook_bytes), read_only=True, data_only=True, keep_vba=True
    )
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Sheet {sheet_name!r} not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]
    target = normalize_compare_text(header)
    for row in ws.iter_rows(min_row=first_row, max_row=last_row):
        for cell in row:
            if normalize_compare_text(cell.value) == target:
                return get_column_letter(cell.column)
    raise KeyError(
        f"Header {header!r} not found in {sheet_name!r} rows {first_row}:{last_row}"
    )


def material_stock_last(
    source_rows: Iterable[Row], destination_codes: Iterable[Any]
) -> dict[str, float | None]:
    """Port Tinh_NVL: sheet #2 B -> H, numeric code only, last duplicate wins.

    Missing codes return None because VBA has CLEAR_IF_NOT_FOUND=True.
    """
    latest: dict[str, float] = {}
    for row in source_rows:
        raw = clean_text(row.get("B")).replace(",", "")
        if not raw:
            continue
        try:
            code = f"{float(raw):.0f}"
        except ValueError:
            continue
        latest[code] = to_number(row.get("H"))

    result: dict[str, float | None] = {}
    for raw in destination_codes:
        value = clean_text(raw).replace(",", "")
        try:
            code = f"{float(value):.0f}" if value else ""
        except ValueError:
            code = ""
        result[normalize_code(raw)] = latest.get(code) if code else None
    return result


def sales_actual_cases(
    source1_rows: Iterable[Row],
    source2_rows: Iterable[Row],
    product_codes: Iterable[Any],
    channel_headers: list[Any],
    divisors: Mapping[str, float],
) -> list[dict[str, Any]]:
    """Port BCBANHANG T:Y values to a shadow result table."""
    totals = aggregate_sales_actual(source1_rows, source2_rows)
    output: list[dict[str, Any]] = []
    for raw_code in product_codes:
        code = normalize_code(raw_code)
        converted = sales_channels_in_cases(
            code, channel_headers, totals, to_number(divisors.get(code))
        )
        row: dict[str, Any] = {"Ma SP": code}
        row.update(converted)
        output.append(row)
    return output
