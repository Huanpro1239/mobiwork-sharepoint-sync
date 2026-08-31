"""Manual image-label preservation across KPI workbook refreshes."""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from urllib.parse import parse_qs, unquote, urlparse

import openpyxl
import pandas as pd

DETAIL_SHEET = "Chi_tiet_Anh_Checkin"
START_ROW = 5
ALLOWED_MANUAL_LABELS = {
    "bien_hieu": "Bien_hieu",
    "trung_bay": "Trung_bay",
    "khong_dat": "Khong_dat",
}


def safe_text(value: object) -> str:
    if value is None:
        return ""
    try:
        if bool(pd.isna(value)):
            return ""
    except (TypeError, ValueError):
        pass
    return str(value)


def _normalise_text(value: object) -> str:
    return " ".join(safe_text(value).strip().casefold().split())


def _normalise_date(value: object) -> str:
    text = safe_text(value).strip()
    parsed = pd.to_datetime(text, errors="coerce")
    return text[:10].casefold() if pd.isna(parsed) else parsed.strftime("%Y-%m-%d")


def _normalise_ordinal(value: object) -> str:
    text = _normalise_text(value).removeprefix("stt").strip()
    try:
        number = float(text)
    except ValueError:
        return text
    return str(int(number)) if number.is_integer() else str(number)


def _filename_from_url(url: str) -> str:
    if not url:
        return ""
    parsed = urlparse(url)
    nested = parse_qs(parsed.query).get("url", ())
    source_path = unquote(nested[0] if nested else parsed.path)
    return PurePosixPath(source_path).name.casefold()


def _keys(
    employee: object,
    when: object,
    customer: object,
    ordinal: object,
    url: object,
) -> tuple[tuple[str, ...], tuple[str, ...]]:
    employee_key = _normalise_text(employee)
    date_key = _normalise_date(when)
    customer_key = _normalise_text(customer)
    ordinal_key = _normalise_ordinal(ordinal)
    filename_key = _filename_from_url(safe_text(url).strip())
    exact = employee_key, date_key, customer_key, ordinal_key, filename_key
    fallback = employee_key, customer_key, ordinal_key, filename_key
    return exact, fallback


def _canonical(value: object, source: Path, row: int) -> str:
    text = safe_text(value).strip()
    canonical = ALLOWED_MANUAL_LABELS.get(text.casefold())
    if canonical is None:
        raise RuntimeError(
            f"Nhãn sửa tay không hợp lệ tại {source}, dòng {row}: {text!r}"
        )
    return canonical


def _register(mapping: dict, key: object, label: str, source: Path, row: int) -> None:
    existing = mapping.get(key)
    if existing is not None and existing != label:
        raise RuntimeError(
            f"Nhãn sửa tay mâu thuẫn tại {source}, dòng {row}: "
            f"{existing!r} và {label!r}"
        )
    mapping[key] = label


@dataclass(frozen=True)
class ManualLabelIndex:
    by_record_id: dict[str, str]
    by_exact_key: dict[tuple[str, ...], str]
    by_fallback_key: dict[tuple[str, ...], str]

    @classmethod
    def empty(cls) -> "ManualLabelIndex":
        return cls({}, {}, {})

    def overlay(self, preferred: "ManualLabelIndex") -> "ManualLabelIndex":
        return ManualLabelIndex(
            {**self.by_record_id, **preferred.by_record_id},
            {**self.by_exact_key, **preferred.by_exact_key},
            {**self.by_fallback_key, **preferred.by_fallback_key},
        )

    def lookup(self, record: dict[str, object]) -> str:
        record_id = safe_text(record.get("record_id")).strip()
        if record_id and record_id in self.by_record_id:
            return self.by_record_id[record_id]
        exact, fallback = _keys(
            record.get("ten_nhan_vien"),
            record.get("ngay"),
            record.get("ma_kh"),
            record.get("stt_hinh"),
            record.get("hinh_anh"),
        )
        return self.by_exact_key.get(exact, self.by_fallback_key.get(fallback, ""))


def labels_from_sheet(sheet, source: Path) -> ManualLabelIndex:
    by_record_id: dict[str, str] = {}
    by_exact_key: dict[tuple[str, ...], str] = {}
    by_fallback_key: dict[tuple[str, ...], str] = {}
    for row in range(START_ROW, sheet.max_row + 1):
        raw_label = safe_text(sheet.cell(row, 8).value).strip()
        if not raw_label:
            continue
        label = _canonical(raw_label, source, row)
        record_id = safe_text(sheet.cell(row, 27).value).strip()
        if record_id:
            _register(by_record_id, record_id, label, source, row)

        hyperlink = sheet.cell(row, 14).hyperlink
        url = hyperlink.target if hyperlink is not None else ""
        exact, fallback = _keys(
            sheet.cell(row, 2).value,
            sheet.cell(row, 3).value,
            sheet.cell(row, 4).value,
            sheet.cell(row, 6).value,
            url,
        )
        if exact[2] and exact[3] and exact[4]:
            _register(by_exact_key, exact, label, source, row)
            _register(by_fallback_key, fallback, label, source, row)
    return ManualLabelIndex(by_record_id, by_exact_key, by_fallback_key)


def load_manual_labels(path: Path) -> ManualLabelIndex:
    if not path.is_file():
        return ManualLabelIndex.empty()
    try:
        workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    except Exception as error:
        raise RuntimeError(
            f"Không đọc được workbook cũ để bảo toàn nhãn sửa tay: "
            f"{type(error).__name__}: {error}"
        ) from error
    try:
        if DETAIL_SHEET not in workbook.sheetnames:
            raise RuntimeError(f"Workbook cũ thiếu sheet {DETAIL_SHEET}; từ chối ghi đè")
        return labels_from_sheet(workbook[DETAIL_SHEET], path)
    finally:
        workbook.close()
