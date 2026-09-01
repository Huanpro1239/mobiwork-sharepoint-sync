"""Fail-closed validation for the approved KPI workbook output contract."""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

SUMMARY_SHEET = "Tong_hop_KPI_Nhan_vien"
CUSTOMER_SHEET = "Chi_tiet_Khach_hang"
DETAIL_SHEET = "Chi_tiet_Anh_Checkin"
ALERT_SHEET = "Canh_bao"
PARAM_SHEET = "Tham_so"
REQUIRED_SHEETS = (SUMMARY_SHEET, CUSTOMER_SHEET, DETAIL_SHEET, ALERT_SHEET, PARAM_SHEET)

SUMMARY_HEADERS = (
    "STT",
    "Tên Nhân Viên",
    "Tổng KH Viếng Thăm",
    "Tổng KH Đạt",
    "KH Chưa Có Đơn",
    "Doanh Số Không Đủ",
    "Không Có Ảnh BH/TB",
    "Ảnh BH/TB Không Đạt",
    "Tỷ Lệ Đạt (Benchmark 50 KH)",
    "Ngày Công Chuẩn",
    "NGÀY CÔNG ĐƯỢC TÍNH",
    "Số KH Mới Đạt",
    "Thưởng Mở Mới (VNĐ)",
    "Số KH Cũ Đạt",
    "Thưởng Chăm Sóc Cũ (VNĐ)",
    "TỔNG TIỀN THƯỞNG KPI (VNĐ)",
    "Đánh Giá / Ghi Chú",
)
CUSTOMER_HEADERS = (
    "STT",
    "Tên Nhân Viên",
    "Mã Khách Hàng",
    "Tên Khách Hàng",
    "Loại Khách Hàng",
    "Số Đơn Hàng",
    "Đơn Lớn Nhất (KTB)",
    "Tổng Sản Lượng (KTB)",
    "Đạt Đơn Hàng",
    "Ghi Tồn",
    "Ảnh Biển Hiệu",
    "Ảnh Trưng Bày",
    "Ảnh Không Đạt",
    "BH Theo Ghi Chú",
    "Đạt Ảnh Check-in",
    "KẾT QUẢ ĐẠT",
    "Chi Tiết / Lý Do Đánh Giá",
    "Dò Ảnh",
)
DETAIL_HEADERS = (
    "STT",
    "Tên Nhân Viên",
    "Ngày",
    "Mã Khách Hàng",
    "Tên Khách Hàng",
    "STT Hình",
    "Phân Loại AI",
    "Nhãn Sửa Tay (Gõ đè vào đây)",
    "Nhãn Dùng Thực Tế",
    "Độ Tin Cậy AI",
    "Tên File",
    "Căn Cứ Nhận Diện",
    "Nội Dung Chữ OCR",
    "Mở Ảnh DMS",
    "Mở File Cục Bộ",
    "Kết Quả Khách Hàng",
    "Trạng Thái Quyết Định",
    "Loại Cảnh",
    "Điểm Scene",
    "Điểm Pass",
    "Điểm Fraud",
    "Độ Tương Đồng Mẫu",
    "3 Tham Chiếu Gần Nhất",
    "Bằng Chứng Detector",
    "Quality Gate",
    "Pipeline Signature",
    "Record ID",
    "Ghi Chú Nguồn",
    "Source Index",
    "Ảnh SHA256",
)
PARAMETERS = (
    ("Ngưỡng KHTC (đơn lớn nhất, KTB)", 3),
    ("Ngưỡng KHĐĐK (cộng dồn tháng, KTB)", 5),
    ("Mức thưởng / KH mới (VNĐ)", 30000),
    ("Mức thưởng / KH cũ (VNĐ)", 10000),
    ("Trần số khách được thưởng / loại", 50),
    ("Trần tổng thưởng / nhân viên (VNĐ)", 4000000),
    ("Benchmark KH (mẫu số tỷ lệ đạt)", 50),
)


def _headers(sheet: Any, count: int) -> tuple[str, ...]:
    return tuple(str(sheet.cell(4, column).value or "").strip() for column in range(1, count + 1))


def _formula_errors(workbook: Any) -> list[str]:
    bad_tokens = ("#REF!", "#DIV/0!", "#NAME?", "#VALUE!")
    errors: list[str] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("=") and any(token in value for token in bad_tokens):
                    errors.append(f"{sheet.title}!{cell.coordinate}={value}")
                    if len(errors) >= 20:
                        return errors
    return errors


def validate_workbook(workbook: Any, *, expected_customers: int | None = None, expected_images: int | None = None) -> dict[str, object]:
    """Validate structure and live-formula wiring before SharePoint can be overwritten."""

    if tuple(workbook.sheetnames) != REQUIRED_SHEETS:
        raise ValueError(f"KPI workbook sheet contract mismatch: {tuple(workbook.sheetnames)!r}")

    checks = (
        (SUMMARY_SHEET, SUMMARY_HEADERS),
        (CUSTOMER_SHEET, CUSTOMER_HEADERS),
        (DETAIL_SHEET, DETAIL_HEADERS),
    )
    for sheet_name, expected in checks:
        actual = _headers(workbook[sheet_name], len(expected))
        if actual != expected:
            for index, (left, right) in enumerate(zip(actual, expected), start=1):
                if left != right:
                    raise ValueError(
                        f"KPI workbook header mismatch {sheet_name}!{openpyxl.utils.get_column_letter(index)}4: "
                        f"{left!r} != {right!r}"
                    )
            raise ValueError(f"KPI workbook header mismatch in {sheet_name}")

    params = workbook[PARAM_SHEET]
    if str(params["A1"].value or "").strip() != "Tham Số Chính Sách & Công Chuẩn":
        raise ValueError("KPI parameter title mismatch")
    for offset, (label, value) in enumerate(PARAMETERS, start=2):
        if str(params.cell(offset, 1).value or "").strip() != label:
            raise ValueError(f"KPI parameter label mismatch at A{offset}")
        actual = params.cell(offset, 2).value
        if float(actual) != float(value):
            raise ValueError(f"KPI parameter value mismatch at B{offset}: {actual!r} != {value!r}")
    if not isinstance(params["B9"].value, (int, float)) or not 20 <= int(params["B9"].value) <= 27:
        raise ValueError(f"KPI working-day value invalid: {params['B9'].value!r}")

    customer = workbook[CUSTOMER_SHEET]
    detail = workbook[DETAIL_SHEET]
    summary = workbook[SUMMARY_SHEET]
    customer_rows = max(0, customer.max_row - 4)
    detail_rows = max(0, detail.max_row - 4)
    summary_rows = max(0, summary.max_row - 4)
    if expected_customers is not None and customer_rows != expected_customers:
        raise ValueError(f"KPI customer row count mismatch: {customer_rows} != {expected_customers}")
    if expected_images is not None and detail_rows != expected_images:
        raise ValueError(f"KPI image row count mismatch: {detail_rows} != {expected_images}")

    if customer_rows:
        required_customer_formulas = ("I5", "K5", "L5", "M5", "O5", "P5", "Q5")
        for coordinate in required_customer_formulas:
            value = customer[coordinate].value
            if not isinstance(value, str) or not value.startswith("="):
                raise ValueError(f"Missing live KPI customer formula at {CUSTOMER_SHEET}!{coordinate}")
    if detail_rows:
        for coordinate in ("I5", "P5"):
            value = detail[coordinate].value
            if not isinstance(value, str) or not value.startswith("="):
                raise ValueError(f"Missing live image formula at {DETAIL_SHEET}!{coordinate}")
    if summary_rows:
        for coordinate in ("C5", "D5", "I5", "K5", "M5", "P5", "Q5"):
            value = summary[coordinate].value
            if not isinstance(value, str) or not value.startswith("="):
                raise ValueError(f"Missing live summary formula at {SUMMARY_SHEET}!{coordinate}")

    formula_errors = _formula_errors(workbook)
    if formula_errors:
        raise ValueError("KPI workbook contains broken formulas: " + "; ".join(formula_errors[:5]))

    alerts = [str(workbook[ALERT_SHEET].cell(row, 1).value or "").strip() for row in range(1, workbook[ALERT_SHEET].max_row + 1)]
    if not any(value.startswith("1. DANH SÁCH DÒNG ĐƠN HÀNG") for value in alerts):
        raise ValueError("KPI alert section 1 is missing")
    if not any(value.startswith("2. DANH SÁCH ẢNH") for value in alerts):
        raise ValueError("KPI alert section 2 is missing")
    if not any(value.startswith("3. BÁO CÁO PHÂN BỔ") for value in alerts):
        raise ValueError("KPI alert section 3 is missing")

    return {
        "valid": True,
        "sheet_count": len(workbook.sheetnames),
        "employee_rows": summary_rows,
        "customer_rows": customer_rows,
        "image_rows": detail_rows,
        "formula_error_count": 0,
    }


def validate_workbook_file(path: str | Path, *, expected_customers: int | None = None, expected_images: int | None = None) -> dict[str, object]:
    workbook = openpyxl.load_workbook(Path(path), data_only=False, read_only=False)
    try:
        return validate_workbook(
            workbook,
            expected_customers=expected_customers,
            expected_images=expected_images,
        )
    finally:
        workbook.close()
