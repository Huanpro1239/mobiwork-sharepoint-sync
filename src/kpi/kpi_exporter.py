"""Template-preserving KPI workbook exporter with live image formulas."""
from __future__ import annotations

import os
from copy import copy
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from rich.console import Console

from kpi.manual_labels import (
    ManualLabelIndex,
    labels_from_sheet,
    load_manual_labels,
    safe_text,
)
from kpi.review_queue import (
    SCOPE_CURRENT_KPI,
    SCOPE_DEFERRED,
    SCOPE_FRAUD_AUDIT,
    SCOPE_HISTORICAL,
    annotate_review_workflow,
    partition_review_rows,
    summarize_review_rows,
)
from kpi.workbook_formulas import (
    START_ROW,
    replace_customer_rows,
    replace_summary_rows,
    update_customer_image_formulas,
    write_parameters,
)
from project_paths import OUTPUT_EXCEL, TEMPLATE_EXCEL

console = Console()
SUMMARY_SHEET = "Tong_hop_KPI_Nhan_vien"
CUSTOMER_SHEET = "Chi_tiet_Khach_hang"
DETAIL_SHEET = "Chi_tiet_Anh_Checkin"
ALERT_SHEET = "Canh_bao"
PARAM_SHEET = "Tham_so"
REQUIRED_SHEETS = (SUMMARY_SHEET, CUSTOMER_SHEET, DETAIL_SHEET, ALERT_SHEET, PARAM_SHEET)

DETAIL_HEADERS = {
    1: "STT",
    2: "Tên Nhân Viên",
    3: "Ngày",
    4: "Mã Khách Hàng",
    5: "Tên Khách Hàng",
    6: "STT Hình",
    7: "Phân Loại AI",
    8: "Nhãn Sửa Tay (Gõ đè vào đây)",
    9: "Nhãn Dùng Thực Tế",
    10: "Độ Tin Cậy AI",
    11: "Tên File",
    12: "Căn Cứ Nhận Diện",
    13: "Nội Dung Chữ OCR",
    14: "Mở Ảnh DMS",
}

AUDIT_COLUMNS = (
    (17, "Trạng Thái Quyết Định", "Trạng Thái Quyết Định"),
    (18, "Loại Cảnh", "Loại Cảnh"),
    (19, "Điểm Scene", "Điểm Scene"),
    (20, "Điểm Pass", "Điểm Pass"),
    (21, "Điểm Fraud", "Điểm Fraud"),
    (22, "Độ Tương Đồng Mẫu", "Độ Tương Đồng Mẫu"),
    (23, "3 Tham Chiếu Gần Nhất", "3 Tham Chiếu Gần Nhất"),
    (24, "Bằng Chứng Detector", "Bằng Chứng Detector"),
    (25, "Quality Gate", "Quality Gate"),
    (26, "Pipeline Signature", "pipeline_signature"),
    (27, "Record ID", "record_id"),
    (28, "Ghi Chú Nguồn", "ghi_chu"),
    (29, "Source Index", "_source_index"),
    (30, "Ảnh SHA256", "image_sha256"),
)
WORKFLOW_COLUMNS = (
    (31, "Nhóm Xử Lý", "_review_scope"),
    (32, "Ưu Tiên Duyệt", "_review_priority"),
    (33, "Hướng Dẫn Xử Lý", "_review_action_reason"),
)
DETAIL_EXTRA_COLUMNS = AUDIT_COLUMNS + WORKFLOW_COLUMNS

_DARK_BLUE = "1F4E78"
_BLUE = "5B9BD5"
_LIGHT_BLUE = "D9EAF7"
_GREEN = "E2F0D9"
_YELLOW = "FFF2CC"
_ORANGE = "FCE4D6"
_RED = "F4CCCC"
_GREY = "E7E6E6"
_DARK_GREY = "7F8C8D"
_TEAL = "DDEBF7"
_WHITE = "FFFFFF"


def _validate_template(workbook) -> None:
    if tuple(workbook.sheetnames) != REQUIRED_SHEETS:
        raise ValueError(
            f"Workbook mẫu không đúng hợp đồng 5 sheet: cần {REQUIRED_SHEETS}, "
            f"nhận {tuple(workbook.sheetnames)}"
        )
    details = workbook[DETAIL_SHEET]
    for column, expected in DETAIL_HEADERS.items():
        actual = safe_text(details.cell(4, column).value).strip()
        if actual != expected:
            coordinate = details.cell(4, column).coordinate
            raise ValueError(
                f"Workbook mẫu sai header {DETAIL_SHEET}!{coordinate}: "
                f"{actual!r}; cần {expected!r}"
            )
    alert_values = [
        safe_text(workbook[ALERT_SHEET].cell(row, 1).value).strip()
        for row in range(1, workbook[ALERT_SHEET].max_row + 1)
    ]
    if not any(value.startswith("2. DANH SÁCH ẢNH") for value in alert_values):
        raise ValueError(f"Workbook mẫu thiếu vùng cảnh báo ảnh trong {ALERT_SHEET}")
    if not any(value.startswith("3. ") for value in alert_values):
        raise ValueError(f"Workbook mẫu thiếu mục 3 trong {ALERT_SHEET}")


def _copy_style(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    target.number_format = source.number_format
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)


def _set_widths(sheet, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _style_header_range(sheet, cell_range: str, fill: str) -> None:
    for row in sheet[cell_range]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(color=_WHITE, bold=True)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )


class KPIExporter:
    def __init__(
        self,
        template_path=TEMPLATE_EXCEL,
        output_path=OUTPUT_EXCEL,
        manual_label_source_path=None,
        announce: bool = True,
    ) -> None:
        self.template_path = Path(template_path)
        self.output_path = Path(output_path)
        self.manual_label_source_path = Path(
            manual_label_source_path if manual_label_source_path is not None else output_path
        )
        self.announce = announce
        self.review_summary: dict[str, object] = {}

    def export_full_workbook(
        self,
        df_results: pd.DataFrame,
        customer_frame: pd.DataFrame,
        period_start,
        kpi_warnings: tuple[str, ...] = (),
        current_pipeline_signature: str = "",
    ) -> Path:
        if not self.template_path.is_file():
            raise FileNotFoundError(f"Không tìm thấy workbook mẫu: {self.template_path}")
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        prior_labels = load_manual_labels(self.manual_label_source_path)
        workbook = openpyxl.load_workbook(self.template_path)
        temporary = self.output_path.with_name(
            f"{self.output_path.stem}_atomic{self.output_path.suffix}"
        )
        try:
            _validate_template(workbook)
            template_labels = labels_from_sheet(workbook[DETAIL_SHEET], self.template_path)
            manual_labels = template_labels.overlay(prior_labels)
            period = pd.Timestamp(period_start).normalize()
            annotated = annotate_review_workflow(
                df_results,
                manual_labels,
                period_start=period,
            )
            self.review_summary = summarize_review_rows(
                df_results,
                manual_labels,
                current_pipeline_signature=current_pipeline_signature,
                period_start=period,
            )
            warning_values = tuple(
                dict.fromkeys((*kpi_warnings, *self.review_summary.get("warnings", [])))
            )
            write_parameters(workbook[PARAM_SHEET], period, warning_values)
            customer_end = replace_customer_rows(workbook[CUSTOMER_SHEET], customer_frame)
            self._replace_detail_rows(
                workbook[DETAIL_SHEET], annotated, manual_labels, customer_end
            )
            update_customer_image_formulas(
                workbook[CUSTOMER_SHEET],
                len(annotated),
                detail_frame=annotated,
            )
            replace_summary_rows(workbook[SUMMARY_SHEET], customer_frame, customer_end)
            self._update_review_alerts(
                workbook[ALERT_SHEET],
                annotated,
                manual_labels,
                period_start=period,
            )
            self._polish_workbook(workbook, period, annotated)
            workbook.calculation.calcMode = "auto"
            workbook.calculation.fullCalcOnLoad = True
            workbook.calculation.forceFullCalc = True
            workbook.save(temporary)
        finally:
            workbook.close()
        os.replace(temporary, self.output_path)
        if self.announce:
            console.print(f"[bold green]Workbook KPI đã xuất: {self.output_path}[/bold green]")
        return self.output_path

    def _replace_detail_rows(self, sheet, frame, manual_labels, customer_end_row: int) -> None:
        max_column = DETAIL_EXTRA_COLUMNS[-1][0]
        style_row = START_ROW if sheet.max_row >= START_ROW else 4
        source_styles = [copy(sheet.cell(style_row, col)._style) for col in range(1, 17)]
        source_formats = [sheet.cell(style_row, col).number_format for col in range(1, 17)]
        header_style = copy(sheet.cell(4, min(14, sheet.max_column))._style)
        if sheet.max_row >= START_ROW:
            sheet.delete_rows(START_ROW, sheet.max_row - START_ROW + 1)
        for column, header, _ in DETAIL_EXTRA_COLUMNS:
            sheet.cell(4, column, header)._style = copy(header_style)
        if sheet.cell(4, 16).value is None:
            sheet.cell(4, 16, "Kết Quả Khách Hàng")._style = copy(header_style)

        records = frame.reset_index(drop=True).to_dict(orient="records")
        lookup_end = max(START_ROW, customer_end_row)
        for offset, record in enumerate(records):
            row = START_ROW + offset
            for col in range(1, 17):
                sheet.cell(row, col)._style = copy(source_styles[col - 1])
                sheet.cell(row, col).number_format = source_formats[col - 1]
            url = safe_text(record.get("hinh_anh"))
            values = {
                1: offset + 1,
                2: record.get("ten_nhan_vien", ""),
                3: record.get("ngay", ""),
                4: record.get("ma_kh", ""),
                5: record.get("ten_kh", ""),
                6: record.get("stt_hinh", ""),
                7: record.get("Phân Loại AI", ""),
                8: manual_labels.lookup(record),
                9: f'=IF(H{row}<>"",H{row},G{row})',
                10: record.get("Độ Tin Cậy AI"),
                11: record.get("Tên File", ""),
                12: record.get("Căn Cứ Nhận Diện", ""),
                13: record.get("Nội Dung Chữ OCR", ""),
                14: "Mở DMS" if url.casefold().startswith(("http://", "https://")) else "",
                15: "",
                16: (
                    f'=IFERROR(VLOOKUP(D{row},\'{CUSTOMER_SHEET}\'!$C$4:$P${lookup_end},14,0),0)'
                ),
            }
            for col, value in values.items():
                sheet.cell(row, col, value)
            if values[14]:
                sheet.cell(row, 14).hyperlink = url
                sheet.cell(row, 14).style = "Hyperlink"
            for column, _, key in DETAIL_EXTRA_COLUMNS:
                target = sheet.cell(row, column, record.get(key, ""))
                reference = sheet.cell(
                    row,
                    12 if column in (23, 24, 26, 27, 28, 30, 33) else 10,
                )
                _copy_style(reference, target)

        end_row = max(START_ROW, START_ROW + len(records) - 1)
        validation = DataValidation(
            type="list", formula1='"Bien_hieu,Trung_bay,Khong_dat"', allow_blank=True
        )
        validation.error = "Chọn một trong ba nhãn KPI chuẩn."
        validation.errorTitle = "Nhãn không hợp lệ"
        validation.prompt = "Chỉ sửa khi đã duyệt ảnh thủ công."
        validation.promptTitle = "Nhãn Sửa Tay"
        sheet.add_data_validation(validation)
        validation.add(f"H{START_ROW}:H{end_row}")
        if records:
            sheet.conditional_formatting.add(
                f"G{START_ROW}:G{end_row}",
                FormulaRule(
                    formula=[f'G{START_ROW}="Can_duyet"'],
                    fill=PatternFill("solid", fgColor=_YELLOW),
                ),
            )
            sheet.conditional_formatting.add(
                f"G{START_ROW}:G{end_row}",
                FormulaRule(
                    formula=[f'G{START_ROW}="Khong_the_cham"'],
                    fill=PatternFill("solid", fgColor=_RED),
                ),
            )
        sheet.freeze_panes = "F5"
        sheet.auto_filter.ref = (
            f"A4:{openpyxl.utils.get_column_letter(max_column)}{end_row}"
        )

    @staticmethod
    def _update_review_alerts(
        sheet,
        frame: pd.DataFrame,
        manual_labels: ManualLabelIndex,
        *,
        period_start=None,
    ) -> None:
        section_row = next(
            (
                row
                for row in range(1, sheet.max_row + 1)
                if safe_text(sheet.cell(row, 1).value).startswith("2. DANH SÁCH ẢNH")
            ),
            None,
        )
        next_section = next(
            (
                row
                for row in range((section_row or 0) + 1, sheet.max_row + 1)
                if safe_text(sheet.cell(row, 1).value).startswith("3. ")
            ),
            None,
        )
        if section_row is None or next_section is None:
            raise ValueError(f"Không tìm thấy đầy đủ mục 2/3 trong sheet {ALERT_SHEET}")
        heading_style = [copy(sheet.cell(section_row, col)._style) for col in range(1, 10)]
        header_source_row = min(section_row + 1, sheet.max_row)
        header_style = [
            copy(sheet.cell(header_source_row, col)._style) for col in range(1, 10)
        ]
        data_source_row = min(section_row + 2, sheet.max_row)
        data_style = [
            copy(sheet.cell(data_source_row, col)._style) for col in range(1, 10)
        ]
        if next_section > section_row + 1:
            sheet.delete_rows(section_row + 1, next_section - section_row - 1)

        partitions = partition_review_rows(
            frame,
            manual_labels,
            period_start=period_start,
        )
        pending_limit = max(
            1,
            int(os.environ.get("KPI_PENDING_ALERT_SAMPLE_LIMIT", "100")),
        )
        sections = (
            (
                "2A. CẦN DUYỆT KPI THÁNG HIỆN TẠI "
                f"({len(partitions.manual_required):,})",
                partitions.manual_required,
                "Không có ảnh nào đang chờ duyệt để xác định KPI tháng hiện tại.",
                _YELLOW,
            ),
            (
                "2B. AUDIT NGHI GIAN LẬN / ĐỐI PHÓ "
                f"({len(partitions.fraud_audit):,})",
                partitions.fraud_audit,
                "Không có ảnh nghi gian lận đang chờ audit.",
                _RED,
            ),
            (
                "2C. ĐÃ ĐỦ BẰNG CHỨNG — CÓ THỂ ĐỂ SAU "
                f"({len(partitions.deferred_review):,})",
                partitions.deferred_review,
                "Không có review dư thừa theo KPI.",
                _LIGHT_BLUE,
            ),
            (
                "2D. REVIEW LỊCH SỬ — KHÔNG CHẶN KPI THÁNG HIỆN TẠI "
                f"({len(partitions.historical_review):,})",
                partitions.historical_review,
                "Không có review lịch sử tồn đọng.",
                _GREY,
            ),
            (
                "2E. LỖI KỸ THUẬT — KHÔNG ĐƯỢC SỬA NHÃN "
                f"({len(partitions.technical):,})",
                partitions.technical,
                "Không có lỗi kỹ thuật bị chặn.",
                _ORANGE,
            ),
            (
                "2F. CHỜ CHẤM AI — HỆ THỐNG TỰ XỬ LÝ "
                f"({len(partitions.pending):,}; hiển thị tối đa {pending_limit:,})",
                partitions.pending.head(pending_limit),
                "Không còn ảnh chờ chấm AI.",
                _TEAL,
            ),
        )
        headers = (
            "STT",
            "Tên Nhân Viên",
            "Mã KH",
            "Tên KH",
            "Nhãn AI",
            "Độ Tin Cậy",
            "Lý Do / Hướng Dẫn",
            "Tên File",
            "Trạng Thái",
        )
        insert_at = section_row + 1
        total_rows = sum(2 + max(1, len(data)) for _, data, _, _ in sections)
        sheet.insert_rows(insert_at, amount=total_rows)

        cursor = insert_at
        for title, data, empty_message, section_fill in sections:
            for col in range(1, 10):
                cell = sheet.cell(cursor, col)
                cell._style = copy(heading_style[col - 1])
                cell.fill = PatternFill("solid", fgColor=section_fill)
                cell.font = Font(bold=True, color="000000")
            sheet.cell(cursor, 1, title)
            cursor += 1

            for col, header in enumerate(headers, start=1):
                target = sheet.cell(cursor, col, header)
                target._style = copy(header_style[col - 1])
            cursor += 1

            if data.empty:
                for col in range(1, 10):
                    sheet.cell(cursor, col)._style = copy(data_style[col - 1])
                sheet.cell(cursor, 1, empty_message)
                cursor += 1
                continue

            for offset, record in enumerate(data.to_dict(orient="records")):
                row = cursor + offset
                values = (
                    offset + 1,
                    record.get("ten_nhan_vien", ""),
                    record.get("ma_kh", ""),
                    record.get("ten_kh", ""),
                    record.get("Phân Loại AI", ""),
                    record.get("Độ Tin Cậy AI"),
                    record.get("_review_action_reason")
                    or record.get("Căn Cứ Nhận Diện", ""),
                    record.get("Tên File", ""),
                    record.get("Trạng Thái Quyết Định", ""),
                )
                for col, value in enumerate(values, start=1):
                    target = sheet.cell(row, col, value)
                    target._style = copy(data_style[col - 1])
            cursor += len(data)

    def _polish_workbook(
        self,
        workbook,
        period: pd.Timestamp,
        detail_frame: pd.DataFrame,
    ) -> None:
        """Apply a clean operator-first view without changing the five-sheet contract."""

        summary = workbook[SUMMARY_SHEET]
        customer = workbook[CUSTOMER_SHEET]
        detail = workbook[DETAIL_SHEET]
        alerts = workbook[ALERT_SHEET]
        params = workbook[PARAM_SHEET]

        for sheet in (summary, customer, detail, alerts, params):
            sheet.sheet_view.showGridLines = False

        summary.sheet_properties.tabColor = _DARK_BLUE
        customer.sheet_properties.tabColor = _BLUE
        detail.sheet_properties.tabColor = "ED7D31"
        alerts.sheet_properties.tabColor = "C00000"
        params.sheet_properties.tabColor = _DARK_GREY

        required = int(self.review_summary.get("manual_review_required_count", 0) or 0)
        fraud = int(self.review_summary.get("fraud_audit_required_count", 0) or 0)
        historical = int(self.review_summary.get("historical_review_count", 0) or 0)
        pending = int(self.review_summary.get("pending_score_count", 0) or 0)
        summary["A2"] = (
            f"KPI {period:%m/%Y} | Cần duyệt KPI: {required:,} | "
            f"Audit fraud: {fraud:,} | Review lịch sử: {historical:,} | "
            f"Chờ AI: {pending:,}"
        )
        summary["A2"].font = Font(bold=True, color=_DARK_BLUE)
        summary["A2"].alignment = Alignment(vertical="center", wrap_text=True)
        summary.row_dimensions[2].height = 30
        summary.freeze_panes = "A5"
        summary.row_dimensions[4].height = 36
        _style_header_range(summary, "A4:Q4", _DARK_BLUE)
        _set_widths(
            summary,
            {
                "A": 6, "B": 24, "C": 15, "D": 14, "E": 14, "F": 16,
                "G": 17, "H": 17, "I": 18, "J": 14, "K": 17, "L": 14,
                "M": 19, "N": 14, "O": 20, "P": 22, "Q": 42,
            },
        )
        summary_end = max(START_ROW, summary.max_row)
        for row in range(START_ROW, summary_end + 1):
            summary.cell(row, 9).number_format = "0.0%"
            for column in (13, 15, 16):
                summary.cell(row, column).number_format = '#,##0 "VNĐ"'

        customer.freeze_panes = "E5"
        customer.row_dimensions[4].height = 36
        _style_header_range(customer, "A4:R4", _BLUE)
        _set_widths(
            customer,
            {
                "A": 6, "B": 22, "C": 16, "D": 32, "E": 16, "F": 12,
                "G": 16, "H": 18, "I": 14, "J": 15, "K": 13, "L": 13,
                "M": 13, "N": 14, "O": 17, "P": 15, "Q": 46, "R": 12,
            },
        )

        detail.row_dimensions[4].height = 40
        _style_header_range(detail, "A4:G4", _DARK_BLUE)
        _style_header_range(detail, "H4:I4", "C55A11")
        _style_header_range(detail, "J4:R4", _BLUE)
        _style_header_range(detail, "S4:AD4", _DARK_GREY)
        _style_header_range(detail, "AE4:AG4", "2F75B5")
        _set_widths(
            detail,
            {
                "A": 6, "B": 22, "C": 12, "D": 16, "E": 32, "F": 8,
                "G": 15, "H": 24, "I": 17, "J": 12, "K": 24, "L": 55,
                "M": 28, "N": 12, "O": 3, "P": 16, "Q": 24, "R": 14,
                "AE": 22, "AF": 14, "AG": 58,
            },
        )
        for column in range(19, 31):
            detail.column_dimensions[openpyxl.utils.get_column_letter(column)].hidden = True
        detail.freeze_panes = "F5"
        detail_end = max(START_ROW, START_ROW + len(detail_frame) - 1)
        detail.auto_filter.ref = f"A4:AG{detail_end}"
        for row in range(START_ROW, detail_end + 1):
            detail.cell(row, 8).fill = PatternFill("solid", fgColor=_YELLOW)
            detail.cell(row, 8).alignment = Alignment(vertical="top", wrap_text=True)
            detail.cell(row, 33).alignment = Alignment(vertical="top", wrap_text=True)
        if len(detail_frame):
            detail.conditional_formatting.add(
                f"G{START_ROW}:G{detail_end}",
                FormulaRule(
                    formula=[f'OR(G{START_ROW}="Bien_hieu",G{START_ROW}="Trung_bay")'],
                    fill=PatternFill("solid", fgColor=_GREEN),
                ),
            )
            detail.conditional_formatting.add(
                f"G{START_ROW}:G{detail_end}",
                FormulaRule(
                    formula=[f'G{START_ROW}="Khong_dat"'],
                    fill=PatternFill("solid", fgColor=_RED),
                ),
            )
            detail.conditional_formatting.add(
                f"AE{START_ROW}:AE{detail_end}",
                FormulaRule(
                    formula=[f'AE{START_ROW}="{SCOPE_CURRENT_KPI}"'],
                    fill=PatternFill("solid", fgColor=_YELLOW),
                ),
            )
            detail.conditional_formatting.add(
                f"AE{START_ROW}:AE{detail_end}",
                FormulaRule(
                    formula=[f'AE{START_ROW}="{SCOPE_FRAUD_AUDIT}"'],
                    fill=PatternFill("solid", fgColor=_RED),
                ),
            )
            detail.conditional_formatting.add(
                f"AE{START_ROW}:AE{detail_end}",
                FormulaRule(
                    formula=[f'AE{START_ROW}="{SCOPE_HISTORICAL}"'],
                    fill=PatternFill("solid", fgColor=_GREY),
                ),
            )
            detail.conditional_formatting.add(
                f"AE{START_ROW}:AE{detail_end}",
                FormulaRule(
                    formula=[f'AE{START_ROW}="{SCOPE_DEFERRED}"'],
                    fill=PatternFill("solid", fgColor=_LIGHT_BLUE),
                ),
            )

        alerts.freeze_panes = "A2"
        _set_widths(
            alerts,
            {
                "A": 7, "B": 22, "C": 16, "D": 32, "E": 15,
                "F": 13, "G": 58, "H": 26, "I": 26, "J": 12,
            },
        )
        for row in range(1, alerts.max_row + 1):
            alerts.cell(row, 7).alignment = Alignment(vertical="top", wrap_text=True)

        _set_widths(params, {"A": 42, "B": 18})
