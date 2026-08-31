"""Live Excel formulas for rolling two-month Sales KPI V2.4."""
from __future__ import annotations

from copy import copy

import openpyxl
import pandas as pd

from kpi.kpi_rules import DEFAULT_KPI_POLICY

START_ROW = 5
DETAIL_SHEET = "Chi_tiet_Anh_Checkin"
CUSTOMER_SHEET = "Chi_tiet_Khach_hang"
PARAM_SHEET = "Tham_so"

CUSTOMER_HEADERS = {
    1: "STT",
    2: "Tên Nhân Viên",
    3: "Mã Khách Hàng",
    4: "Tên Khách Hàng",
    5: "Lượt Viếng Thăm Tháng M",
    6: "Loại KH (Mới/Cũ)",
    7: "Đơn Lớn Nhất 2T (KTB)",
    8: "Tổng Sản Lượng 2T (KTB)",
    9: "Phân Loại Doanh Số",
    10: "Ghi Tồn 2T",
    11: "Ảnh Biển Hiệu",
    12: "Ảnh Trưng Bày",
    13: "Ảnh Không Đạt",
    14: "BH Theo Ghi Chú",
    15: "Đủ Ảnh 2T",
    16: "KẾT QUẢ ĐẠT",
    17: "Lý Do",
    18: "Ngày Hoạt Động Đầu Tiên",
    19: "Số Đơn 2T",
    20: "Ghi Chú 2T",
    21: "Kỳ Tính M",
}

SUMMARY_HEADERS = {
    1: "STT",
    2: "Tên Nhân Viên",
    3: "Tổng KH Viếng Thăm M",
    4: "Tổng KH Đạt",
    5: "KH Chưa Có Đơn",
    6: "Doanh Số Không Đủ",
    7: "Không Đạt Ghi Tồn",
    8: "Không Đạt Ảnh BH/TB",
    9: "Tỷ Lệ Đạt (Benchmark)",
    10: "Ngày Công Chuẩn",
    11: "NGÀY CÔNG ĐƯỢC TÍNH",
    12: "Số KH Mới Đạt",
    13: "Thưởng Mở Mới (VNĐ)",
    14: "Số KH Cũ Đạt",
    15: "Thưởng Chăm Sóc Cũ (VNĐ)",
    16: "TỔNG TIỀN THƯỞNG KPI (VNĐ)",
    17: "Đánh Giá / Ghi Chú",
}


def _row_styles(sheet, end_column: int) -> tuple[list, list]:
    source_row = START_ROW if sheet.max_row >= START_ROW else 4
    styles = [copy(sheet.cell(source_row, col)._style) for col in range(1, end_column + 1)]
    formats = [sheet.cell(source_row, col).number_format for col in range(1, end_column + 1)]
    return styles, formats


def _reset_table(sheet, headers: dict[int, str]) -> tuple[list, list]:
    end_column = max(headers)
    styles, formats = _row_styles(sheet, end_column)
    header_style = copy(sheet.cell(4, min(sheet.max_column, end_column))._style)
    if sheet.max_row >= START_ROW:
        sheet.delete_rows(START_ROW, sheet.max_row - START_ROW + 1)
    for column, header in headers.items():
        cell = sheet.cell(4, column, header)
        cell._style = copy(header_style)
    return styles, formats


def write_parameters(sheet, period_start: pd.Timestamp, warnings: tuple[str, ...]) -> None:
    policy = DEFAULT_KPI_POLICY
    sheet["A1"] = "THAM SỐ KPI SALES - GHÉP 2 THÁNG"
    rows = (
        (3, "Kỳ tính (M)", period_start.to_pydatetime()),
        (4, "Benchmark khách hàng", policy.benchmark_customers),
        (5, "Ngưỡng KHTC - đơn đơn lẻ (KTB)", policy.khtc_single_order_ktb),
        (6, "Ngưỡng KHĐĐK - tổng 2 tháng (KTB)", policy.khddk_total_ktb),
        (7, "Thưởng KH mới (VNĐ)", policy.new_customer_reward_vnd),
        (8, "Thưởng KH cũ (VNĐ)", policy.old_customer_reward_vnd),
        (9, "Trần KH thưởng / loại", policy.reward_customer_cap),
        (10, "Trần thưởng KPI (VNĐ)", policy.reward_total_cap_vnd),
    )
    for row, label, value in rows:
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, value)
    sheet["A11"] = "Ngày công chuẩn (trừ Chủ nhật)"
    sheet["B11"] = '=NETWORKDAYS.INTL(B3,EOMONTH(B3,0),"0000001")'
    sheet["A12"] = "Tháng M-1"
    sheet["B12"] = "=EDATE(B3,-1)"
    sheet["B3"].number_format = "dd/mm/yyyy"
    sheet["B12"].number_format = "dd/mm/yyyy"
    for cell in ("B7", "B8", "B10"):
        sheet[cell].number_format = '#,##0 "VNĐ"'
    sheet["A14"] = "CẢNH BÁO DỮ LIỆU"
    for row in range(15, max(25, sheet.max_row + 1)):
        sheet.cell(row, 1, None)
    for offset, warning in enumerate(warnings or ("Không có cảnh báo dữ liệu KPI.",), start=15):
        sheet.cell(offset, 1, warning)
    sheet.column_dimensions["A"].width = max(sheet.column_dimensions["A"].width or 0, 48)
    sheet.column_dimensions["B"].width = max(sheet.column_dimensions["B"].width or 0, 18)


def replace_customer_rows(sheet, frame: pd.DataFrame) -> int:
    styles, formats = _reset_table(sheet, CUSTOMER_HEADERS)
    records = frame.reset_index(drop=True).to_dict(orient="records")
    for offset, record in enumerate(records):
        row = START_ROW + offset
        for col in range(1, max(CUSTOMER_HEADERS) + 1):
            sheet.cell(row, col)._style = copy(styles[col - 1])
            sheet.cell(row, col).number_format = formats[col - 1]
        first_activity = record.get("first_activity_date")
        period = record.get("period_start")
        values = {
            1: offset + 1,
            2: record.get("ten_nhan_vien", ""),
            3: record.get("ma_kh", ""),
            4: record.get("ten_kh", ""),
            5: int(record.get("visit_count_m", 0) or 0),
            6: f'=IF(R{row}="","Không rõ",IF(R{row}<\'{PARAM_SHEET}\'!$B$3,"Cũ","Mới"))',
            7: float(record.get("max_order_2m_ktb", 0) or 0),
            8: float(record.get("total_order_2m_ktb", 0) or 0),
            9: f'=IF(G{row}>=\'{PARAM_SHEET}\'!$B$5,"KHTC",IF(H{row}>=\'{PARAM_SHEET}\'!$B$6,"KHĐĐK","Không đạt DS"))',
            10: 1 if bool(record.get("ghi_ton_2m", False)) else 0,
            11: 0,
            12: 0,
            13: 0,
            14: 1 if bool(record.get("valid_sign_note_2m", False)) else 0,
            15: f'=IF(AND(OR(K{row}>=1,N{row}=1),L{row}>=1),1,0)',
            16: f'=IF(E{row}<=0,"Không xét",IF(I{row}="Không đạt DS","Không Đạt",IF(J{row}<>1,"Không Đạt",IF(O{row}<>1,"Không Đạt",I{row}))))',
            17: f'=IF(P{row}="KHTC","Đạt KHTC",IF(P{row}="KHĐĐK","Đạt KHĐĐK",IF(E{row}<=0,"Không có viếng thăm tháng M",IF(I{row}="Không đạt DS","Doanh số 2 tháng không đủ",IF(J{row}<>1,"Thiếu ghi tồn M-1/M",IF(O{row}<>1,"Thiếu Biển hiệu/Trưng bày M-1/M",""))))))',
            19: int(record.get("order_count_2m", 0) or 0),
            20: record.get("ghi_chu_2m", ""),
        }
        for col, value in values.items():
            sheet.cell(row, col, value)
        if first_activity is not None and not pd.isna(first_activity):
            sheet.cell(row, 18, pd.Timestamp(first_activity).to_pydatetime())
            sheet.cell(row, 18).number_format = "dd/mm/yyyy"
        if period is not None and not pd.isna(period):
            sheet.cell(row, 21, pd.Timestamp(period).to_pydatetime())
            sheet.cell(row, 21).number_format = "mm/yyyy"
        sheet.cell(row, 7).number_format = "0.00"
        sheet.cell(row, 8).number_format = "0.00"
    end_row = START_ROW + len(records) - 1 if records else START_ROW - 1
    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A4:{openpyxl.utils.get_column_letter(max(CUSTOMER_HEADERS))}{max(4, end_row)}"
    return end_row


def update_customer_image_formulas(sheet, detail_count: int) -> None:
    detail_end = max(START_ROW, START_ROW + detail_count - 1)
    for row in range(START_ROW, sheet.max_row + 1):
        customer = f"C{row}"
        base = (
            f"'{DETAIL_SHEET}'!$D$5:$D${detail_end},{customer},"
            f"'{DETAIL_SHEET}'!$I$5:$I${detail_end}"
        )
        sheet[f"K{row}"] = f'=COUNTIFS({base},"Bien_hieu")'
        sheet[f"L{row}"] = f'=COUNTIFS({base},"Trung_bay")'
        sheet[f"M{row}"] = f'=COUNTIFS({base},"Khong_dat")'


def replace_summary_rows(sheet, customer_frame: pd.DataFrame, customer_end_row: int) -> None:
    styles, formats = _reset_table(sheet, SUMMARY_HEADERS)
    employees = []
    if not customer_frame.empty and "ten_nhan_vien" in customer_frame.columns:
        employees = [
            str(value)
            for value in customer_frame["ten_nhan_vien"].dropna().drop_duplicates().tolist()
            if str(value).strip()
        ]
    customer_end = max(START_ROW, customer_end_row)
    for offset, employee in enumerate(employees):
        row = START_ROW + offset
        for col in range(1, max(SUMMARY_HEADERS) + 1):
            sheet.cell(row, col)._style = copy(styles[col - 1])
            sheet.cell(row, col).number_format = formats[col - 1]
        ranges = {
            key: f"'{CUSTOMER_SHEET}'!${key}$5:${key}${customer_end}"
            for key in ("B", "E", "F", "H", "I", "J", "O", "P")
        }
        b, e, f, h, i, j, o, p = (ranges[key] for key in ("B", "E", "F", "H", "I", "J", "O", "P"))
        sheet.cell(row, 1, offset + 1)
        sheet.cell(row, 2, employee)
        sheet.cell(row, 3, f'=COUNTIFS({b},B{row},{e},">0")')
        sheet.cell(row, 4, f'=COUNTIFS({b},B{row},{p},"KHTC")+COUNTIFS({b},B{row},{p},"KHĐĐK")')
        sheet.cell(row, 5, f'=COUNTIFS({b},B{row},{h},0)')
        sheet.cell(row, 6, f'=COUNTIFS({b},B{row},{i},"Không đạt DS")')
        sheet.cell(row, 7, f'=COUNTIFS({b},B{row},{j},0)')
        sheet.cell(row, 8, f'=COUNTIFS({b},B{row},{o},0)')
        sheet.cell(row, 9, f'=MIN(D{row}/\'{PARAM_SHEET}\'!$B$4,1)')
        sheet.cell(row, 10, f'=\'{PARAM_SHEET}\'!$B$11')
        sheet.cell(row, 11, f'=I{row}*J{row}')
        new_pass = f'COUNTIFS({b},B{row},{f},"Mới",{p},"KHTC")+COUNTIFS({b},B{row},{f},"Mới",{p},"KHĐĐK")'
        old_pass = f'COUNTIFS({b},B{row},{f},"Cũ",{p},"KHTC")+COUNTIFS({b},B{row},{f},"Cũ",{p},"KHĐĐK")'
        sheet.cell(row, 12, f'={new_pass}')
        sheet.cell(row, 13, f'=MIN(L{row},\'{PARAM_SHEET}\'!$B$9)*\'{PARAM_SHEET}\'!$B$7')
        sheet.cell(row, 14, f'={old_pass}')
        sheet.cell(row, 15, f'=MIN(N{row},\'{PARAM_SHEET}\'!$B$9)*\'{PARAM_SHEET}\'!$B$8')
        sheet.cell(row, 16, f'=MIN(M{row}+O{row},\'{PARAM_SHEET}\'!$B$10)')
        sheet.cell(row, 17, f'=IF(C{row}=0,"Không có KH viếng thăm",IF(D{row}=0,"Chưa có KH đạt",""))')
        sheet.cell(row, 9).number_format = "0.00%"
        sheet.cell(row, 11).number_format = "0.00"
        for col in (13, 15, 16):
            sheet.cell(row, col).number_format = '#,##0 "VNĐ"'
    end_row = START_ROW + len(employees) - 1 if employees else START_ROW - 1
    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A4:Q{max(4, end_row)}"
